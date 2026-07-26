#!/usr/bin/env python3
"""Materialize and apply the user-authorized PENC276 CNBE32 encoding batch.

The 276 rows use human-approved structure labels, retained v1.1 stroke/index
values, and Kangxi radical numbers from the frozen inventory.  The command is
dry-run by default; ``--apply`` is required for source and SQLite writes.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import shutil
import sqlite3
from datetime import datetime, timezone
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
INVENTORY = ROOT / "evidence/8105/PENDING_276_ENCODING_INVENTORY.csv"
STRUCTURES = ROOT / "review_packets/8105_full/8105_hanzi_decomp_v03_human_approved_structure_packet.csv"
T3_LEDGER = ROOT / "evidence/8105/pending276/T3_169_276_HUMAN_DECOMPOSITION_LEDGER.csv"
RADICAL_NAMES = ROOT / "evidence/8105/pending276/radical_name_to_kangxi.json"
CANDIDATES = ROOT / "evidence/8105/pending276/PENC276_AUTHORIZED_ENCODING_CANDIDATES.csv"
REPORT_JSON = ROOT / "reports/PENC276_AUTHORIZED_ENCODING_APPLY.json"
REPORT_MD = ROOT / "reports/PENC276_AUTHORIZED_ENCODING_APPLY.md"
RUNTIME_JSON = ROOT / "data/cnbe32.json"
ROOT_DB = ROOT / "data/cnbe32.db"
PACKAGE_DB = ROOT / "src/cnbe32/data/cnbe32.db"
STRUCT_CODES = {name: code for code, name in enumerate(["独体字", "上下", "上中下", "左右", "左中右", "左上包", "右上包", "左三包", "左下包", "上三包", "下三包", "全包围", "镶嵌"])}
CREATE_SQL = """CREATE TABLE cnbe32 (unicode INTEGER PRIMARY KEY,char TEXT,cnbe INTEGER,radix INTEGER,radix_name TEXT,strokes INTEGER,struct_type INTEGER,struct_name TEXT,idx INTEGER,track TEXT,needs_encoding INTEGER DEFAULT 0);
CREATE INDEX idx_cnbe ON cnbe32(cnbe); CREATE INDEX idx_radix ON cnbe32(radix); CREATE INDEX idx_strokes ON cnbe32(strokes);
CREATE TABLE migration_meta (version TEXT, applied_at TEXT, script_version TEXT, updates INT, inserts INT, backup TEXT);"""


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8", newline="") as stream:
        return list(csv.DictReader(stream))


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def code(radix: int, strokes: int, struct_type: int, index: int) -> int:
    return (radix << 24) | (strokes << 19) | (struct_type << 15) | (index << 4)


def decode(value: int) -> tuple[int, int, int, int, int]:
    return ((value >> 24) & 0xFF, (value >> 19) & 0x1F, (value >> 15) & 0x0F, (value >> 4) & 0x7FF, value & 0x0F)


def radical_names() -> dict[int, str]:
    names = json.loads(RADICAL_NAMES.read_text(encoding="utf-8"))["mapping"]
    preferred = {code: name for code, name in {
        8:"亠",9:"人",17:"凵",30:"口",72:"日",95:"王",96:"玄",169:"門",170:"阜",211:"齒",212:"龍"
    }.items()}
    for name, values in names.items():
        for value in values:
            preferred.setdefault(int(value), name)
    return preferred


def read_review_decompositions(workbook: Path) -> dict[str, str]:
    from openpyxl import load_workbook  # Local administrator import only.
    book = load_workbook(workbook, read_only=True, data_only=True)
    sheet = book[book.sheetnames[0]]
    values = sheet.iter_rows(values_only=True)
    headers = next(values)
    rows = [dict(zip(headers, row)) for row in values]
    required = {"清单编号", "审核拆解或部件（填写）", "审核结论（填写）"}
    if len(rows) != 168 or not required <= set(headers):
        raise ValueError("completed reviewer workbook does not match the 168-row PENC276 schema")
    if any(str(row["审核结论（填写）"]).strip() != "通过" for row in rows):
        raise ValueError("completed reviewer workbook contains a non-pass decision")
    if any(not str(row["审核拆解或部件（填写）"] or "").strip() for row in rows):
        raise ValueError("completed reviewer workbook contains an empty decomposition")
    return {str(row["清单编号"]): str(row["审核拆解或部件（填写）"]).strip() for row in rows}


def materialize(workbook: Path) -> list[dict[str, str]]:
    inventory = read_csv(INVENTORY)
    if len(inventory) != 276 or len({row["row_id"] for row in inventory}) != 276:
        raise ValueError("PENC276 inventory is not unique 276 rows")
    structures = {row["character"]: row for row in read_csv(STRUCTURES)}
    t3 = {row["row_id"]: row for row in read_csv(T3_LEDGER)}
    review = read_review_decompositions(workbook)
    names = radical_names()
    candidates: list[dict[str, str]] = []
    for row in inventory:
        approved = structures.get(row["char"])
        if not approved or approved["human_review_status"] != "HUMAN_APPROVED_AGENT_STRUCTURE_CANDIDATE":
            raise ValueError(f"missing approved structure for {row['row_id']}")
        structure = approved["candidate_structure_label"]
        if structure not in STRUCT_CODES or int(approved["agent_struct_type"]) != STRUCT_CODES[structure]:
            raise ValueError(f"invalid approved structure for {row['row_id']}")
        radical, strokes, index = int(row["unihan_kangxi_radical"]), int(row["strokes_db"]), int(row["idx"])
        if not (1 <= radical <= 214 and 1 <= strokes <= 31 and 0 <= index <= 2047):
            raise ValueError(f"bitfield range failure for {row['row_id']}")
        value = code(radical, strokes, STRUCT_CODES[structure], index)
        if decode(value) != (radical, strokes, STRUCT_CODES[structure], index, 0):
            raise ValueError(f"round-trip failure for {row['row_id']}")
        decomposition = review.get(row["row_id"], "")
        if not decomposition:
            ledger = t3.get(row["row_id"])
            decomposition = "".join((ledger or {}).get(key, "") for key in ("component_1", "component_2"))
        if not decomposition:
            raise ValueError(f"missing human decomposition for {row['row_id']}")
        candidates.append({
            "row_id":row["row_id"], "char":row["char"], "unicode":row["unicode_dec"], "cnbe":str(value), "cnbe_hex":f"0x{value:08X}",
            "radix":str(radical), "radix_name":names[radical], "strokes":str(strokes), "struct_type":str(STRUCT_CODES[structure]),
            "struct_name":structure, "idx":str(index), "ext":"0", "human_decomposition":decomposition,
            "authority":"HUMAN_AUDIT_PROJECT_BASELINE_USER_AUTHORIZED_2026_07_27",
        })
    return candidates


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=list(rows[0]), lineterminator="\n")
        writer.writeheader(); writer.writerows(rows)


def rebuild(path: Path, rows: list[dict]) -> dict[str, object]:
    temporary = path.with_suffix(".tmp.db")
    if temporary.exists(): temporary.unlink()
    with sqlite3.connect(temporary) as connection:
        connection.executescript(CREATE_SQL)
        connection.executemany("INSERT INTO cnbe32 VALUES (?,?,?,?,?,?,?,?,?,?,?)", [
            (r["unicode"],r["char"],r["cnbe"],r["radix"],r["radix_name"],r["strokes"],r["struct_type"],r["struct_name"],r["index"],r.get("track","standard"),r.get("needs_encoding",0)) for r in rows])
        connection.execute("INSERT INTO migration_meta VALUES (?,?,?,?,?,?)", ("v1.2-penc276",datetime.now(timezone.utc).isoformat(),"1.0.0",276,276,"build/penc276-authorized-backup"))
        integrity = connection.execute("PRAGMA integrity_check").fetchone()[0]
        count = connection.execute("SELECT COUNT(*) FROM cnbe32").fetchone()[0]
    temporary.replace(path)
    return {"path":str(path.relative_to(ROOT)),"rows":count,"integrity":integrity,"sha256":sha256(path)}


def database_baseline(path: Path) -> list[dict]:
    """Read the v1.1 database without normalizing non-target records."""
    columns = "unicode,char,cnbe,radix,radix_name,strokes,struct_type,struct_name,idx,track,needs_encoding"
    with sqlite3.connect(path) as connection:
        records = connection.execute(f"SELECT {columns} FROM cnbe32 ORDER BY unicode").fetchall()
    if len(records) != 21178:
        raise ValueError("unexpected v1.1 database baseline")
    return [
        {
            "unicode": row[0], "char": row[1], "cnbe": row[2], "radix": row[3],
            "radix_name": row[4], "strokes": row[5], "struct_type": row[6],
            "struct_name": row[7], "index": row[8], "track": row[9],
            "needs_encoding": row[10],
        }
        for row in records
    ]


def copy_backup_once(source: Path, destination: Path) -> None:
    if not destination.exists():
        shutil.copy2(source, destination)


def apply(candidates: list[dict[str, str]], baseline_json: Path, baseline_db: Path) -> dict:
    model=json.loads(baseline_json.read_text(encoding="utf-8")); original=model["characters"]
    if len(original)!=20902 or len({r["unicode"] for r in original})!=20902: raise ValueError("unexpected source JSON baseline")
    baseline_rows = database_baseline(baseline_db)
    additions=[{"char":r["char"],"unicode":int(r["unicode"]),"cnbe":int(r["cnbe"]),"radix":int(r["radix"]),"radix_name":r["radix_name"],"strokes":int(r["strokes"]),"struct_type":int(r["struct_type"]),"struct_name":r["struct_name"],"index":int(r["idx"]),"track":"standard","needs_encoding":0} for r in candidates]
    if set(r["unicode"] for r in original) & set(r["unicode"] for r in additions): raise ValueError("candidate already exists in source JSON")
    rows=original+additions
    updates = {int(row["unicode"]): row for row in additions}
    if not set(updates) <= {row["unicode"] for row in baseline_rows}:
        raise ValueError("candidate is absent from database baseline")
    # Preserve every non-target database value exactly; this migration owns 276 rows only.
    database_rows = [dict(row, **updates[row["unicode"]]) if row["unicode"] in updates else row for row in baseline_rows]
    model["characters"]=rows; model.setdefault("metadata",{}).update({"total":21178,"penc276_authorized_encoding_rows":276,"penc276_authority":"human_audit_project_baseline_user_authorized_2026_07_27"})
    backup=ROOT/"build/penc276-authorized-backup"; backup.mkdir(parents=True,exist_ok=True)
    copy_backup_once(RUNTIME_JSON, backup / "cnbe32.json")
    copy_backup_once(ROOT_DB, backup / "root_cnbe32.db")
    copy_backup_once(PACKAGE_DB, backup / "package_cnbe32.db")
    RUNTIME_JSON.write_text(json.dumps(model,ensure_ascii=False,indent=2,sort_keys=True)+"\n",encoding="utf-8")
    databases=[rebuild(ROOT_DB,database_rows),rebuild(PACKAGE_DB,database_rows)]
    return {"runtime_rows":len(rows),"candidate_rows":len(candidates),"databases":databases,"backup":str(backup.relative_to(ROOT))}


def main() -> None:
    parser=argparse.ArgumentParser(); parser.add_argument("--review-xlsx",type=Path,required=True); parser.add_argument("--apply",action="store_true")
    parser.add_argument("--baseline-json", type=Path, default=RUNTIME_JSON)
    parser.add_argument("--baseline-db", type=Path, default=ROOT_DB)
    args=parser.parse_args(); candidates=materialize(args.review_xlsx); write_csv(CANDIDATES,candidates)
    result={"status":"DRY_RUN_READY","candidate_rows":len(candidates),"candidate_sha256":sha256(CANDIDATES),"source_table_write_authorized":False}
    if args.apply:
        result.update(apply(candidates, args.baseline_json, args.baseline_db)); result.update({"status":"PASS_PENC276_AUTHORIZED_ENCODING_APPLIED","source_table_write_authorized":True})
    REPORT_JSON.write_text(json.dumps(result,ensure_ascii=False,indent=2)+"\n",encoding="utf-8")
    REPORT_MD.write_text(f"# PENC276 授权 CNBE 编码报告\n\n- 状态：`{result['status']}`\n- 候选行：`{len(candidates)}`\n- 候选 SHA-256：`{result['candidate_sha256']}`\n- 源表写入：`{result['source_table_write_authorized']}`\n",encoding="utf-8")
    print(result["status"])

if __name__ == "__main__": main()
