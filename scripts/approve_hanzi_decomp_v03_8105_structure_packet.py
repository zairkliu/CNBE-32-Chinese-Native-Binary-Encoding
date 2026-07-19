#!/usr/bin/env python3
"""Materialize the human-approved 8105 v0.3 Agent structure packet.

This script applies the explicit human review decision:

- hanzi_decomp_v0.3 is accepted as the current Agent candidate reference;
- the only remaining blank row, 孓 (U+5B53), is an independent character;
- output remains an Agent candidate layer, not a CNBE32/database writer.
"""

from __future__ import annotations

import csv
import json
import sys
from collections import Counter
from pathlib import Path
from typing import Any

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from scripts.run_hanzi_decomp_v03_8105_adapter import APPROVED_STRUCTURES

BASE_PACKET = Path("review_packets/8105_full/8105_hanzi_decomp_v03_structure_code_candidate_packet.csv")
ADAPTER_ALL = Path("review_packets/8105_full/8105_hanzi_decomp_v03_adapter_all.csv")

OUTPUT_DIR = Path("review_packets/8105_full")
APPROVED_PACKET = OUTPUT_DIR / "8105_hanzi_decomp_v03_human_approved_structure_packet.csv"
APPROVED_CHANGES = OUTPUT_DIR / "8105_hanzi_decomp_v03_human_approved_changes.csv"
APPROVED_XLSX = Path("outputs/8105_hanzi_decomp_v03_human_approved_structure_review.xlsx")

JSON_REPORT = Path("reports/8105_hanzi_decomp_v03_human_approved_structure_packet.json")
MD_REPORT = Path("reports/8105_HANZI_DECOMP_V03_HUMAN_APPROVED_STRUCTURE_PACKET.md")

EXPECTED_ROWS = 8105
HUMAN_DECISION_ID = "HUMAN_REVIEW_2026_07_19_V03_ACCEPTED_JUE_DUTIZI"
JUE_CHAR = "孓"
JUE_UNICODE = "U+5B53"


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames: list[str] = []
    for row in rows:
        for key in row:
            if key not in fieldnames:
                fieldnames.append(key)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def apply_human_decision(row: dict[str, Any], adapter: dict[str, str] | None) -> tuple[dict[str, Any], bool]:
    out = dict(row)
    before = {
        "candidate_structure_label": row.get("candidate_structure_label", ""),
        "candidate_decomposition": row.get("candidate_decomposition", ""),
        "agent_struct_type": row.get("agent_struct_type", ""),
    }
    out["human_review_decision_id"] = HUMAN_DECISION_ID
    out["human_review_decision"] = "V03_ACCEPTED_AS_AGENT_REFERENCE"
    out["human_review_status"] = "HUMAN_APPROVED_AGENT_STRUCTURE_CANDIDATE"
    out["authority_boundary"] = "AGENT_STANDARD_CANDIDATE_NOT_NATIONAL_STANDARD"
    out["final_source_table_write_status"] = "NO_SOURCE_TABLE_WRITE"
    out["final_cnbe32_write_status"] = "NO_CNBE32_WRITE"
    out["final_database_rebuild_status"] = "NO_DATABASE_REBUILD"

    if adapter and adapter.get("tool_structure"):
        out["candidate_structure_label"] = adapter["tool_structure"]
        out["candidate_decomposition"] = adapter.get("tool_decomposition", "")
        out["direct_component_candidates"] = adapter.get("tool_top_level_parts", "")
        out["agent_struct_type"] = adapter.get("tool_struct_type", "")
        out["human_review_basis"] = "HUMAN_CONFIRMED_HANZI_DECOMP_V03"
        out["candidate_structure_status"] = "HUMAN_APPROVED_V03_AGENT_REFERENCE"
        out["candidate_decomposition_status"] = "HUMAN_APPROVED_V03_AGENT_REFERENCE"
        out["standardizer_status"] = "HUMAN_APPROVED_V03_AGENT_REFERENCE_NO_SOURCE_WRITE"
        out["review_status"] = "HUMAN_APPROVED_FOR_AGENT_CANDIDATE_LAYER"
    elif out.get("character") == JUE_CHAR and out.get("unicode_codepoint") == JUE_UNICODE:
        out["candidate_structure_label"] = "独体字"
        out["candidate_decomposition"] = JUE_CHAR
        out["direct_component_candidates"] = JUE_CHAR
        out["agent_struct_type"] = str(APPROVED_STRUCTURES["独体字"])
        out["human_review_basis"] = "HUMAN_CONFIRMED_JUE_U5B53_DUTIZI"
        out["candidate_structure_status"] = "HUMAN_APPROVED_INDEPENDENT_CHARACTER"
        out["candidate_decomposition_status"] = "HUMAN_APPROVED_INDEPENDENT_CHARACTER"
        out["standardizer_status"] = "HUMAN_APPROVED_REMAINING_BLANK_FILLED_NO_SOURCE_WRITE"
        out["review_status"] = "HUMAN_APPROVED_FOR_AGENT_CANDIDATE_LAYER"
    else:
        out["human_review_basis"] = "HUMAN_CONFIRMED_EXISTING_CANDIDATE_WHEN_V03_TOOL_GAP"
        out["review_status"] = "HUMAN_APPROVED_FOR_AGENT_CANDIDATE_LAYER"

    after = {
        "candidate_structure_label": out.get("candidate_structure_label", ""),
        "candidate_decomposition": out.get("candidate_decomposition", ""),
        "agent_struct_type": out.get("agent_struct_type", ""),
    }
    return out, before != after


def build() -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    base_rows = read_csv(BASE_PACKET)
    adapter_rows = {row["character"]: row for row in read_csv(ADAPTER_ALL)}
    approved: list[dict[str, Any]] = []
    changed: list[dict[str, Any]] = []
    for row in base_rows:
        out, did_change = apply_human_decision(row, adapter_rows.get(row["character"]))
        approved.append(out)
        if did_change:
            changed.append(out)

    structure_values = {row["candidate_structure_label"] for row in approved}
    bad_structures = sorted(value for value in structure_values if value not in APPROVED_STRUCTURES)
    duplicate_chars = [
        char
        for char, count in Counter(row["character"] for row in approved).items()
        if count != 1
    ]
    duplicate_unicode = [
        code
        for code, count in Counter(row["unicode_codepoint"] for row in approved).items()
        if count != 1
    ]
    by_char = {row["character"]: row for row in approved}
    status_counts = Counter(row["human_review_basis"] for row in approved)
    structure_counts = Counter(row["candidate_structure_label"] for row in approved)
    checks = {
        "row_count_is_8105": len(approved) == EXPECTED_ROWS,
        "unique_character_count_is_8105": len({row["character"] for row in approved}) == EXPECTED_ROWS,
        "unique_unicode_count_is_8105": len({row["unicode_codepoint"] for row in approved}) == EXPECTED_ROWS,
        "all_structures_nonblank": all(row["candidate_structure_label"] for row in approved),
        "all_structures_allowed": not bad_structures,
        "all_structures_have_agent_code": all(row["agent_struct_type"] != "" for row in approved),
        "jue_is_independent": by_char[JUE_CHAR]["candidate_structure_label"] == "独体字"
        and by_char[JUE_CHAR]["agent_struct_type"] == "0"
        and by_char[JUE_CHAR]["candidate_decomposition"] == JUE_CHAR,
        "known_legacy_regressions_are_left_right": all(
            by_char[char]["candidate_structure_label"] == "左右"
            and by_char[char]["agent_struct_type"] == "3"
            for char in ("侵", "偶", "冁")
        ),
        "no_source_table_writes": all(row["final_source_table_write_status"] == "NO_SOURCE_TABLE_WRITE" for row in approved),
        "no_cnbe32_writes": all(row["final_cnbe32_write_status"] == "NO_CNBE32_WRITE" for row in approved),
        "no_database_rebuild": all(row["final_database_rebuild_status"] == "NO_DATABASE_REBUILD" for row in approved),
    }
    report = {
        "report_schema_version": "1.0",
        "mode": "hanzi_decomp_v03_human_approved_8105_agent_structure_packet",
        "overall_status": "PASS_HUMAN_APPROVED_V03_8105_AGENT_STRUCTURE_PACKET_READY"
        if all(checks.values())
        else "BLOCKED",
        "human_decision": {
            "decision_id": HUMAN_DECISION_ID,
            "decision_text": "孓为独体，其余审核通过，0.3为准",
            "jue_char": JUE_CHAR,
            "jue_unicode": JUE_UNICODE,
            "v03_authorized_as": "agent_standard_candidate_reference",
            "not_authorized_as": [
                "national_standard_output",
                "source_table_write",
                "cnbe32_write",
                "sqlite_database_rebuild",
            ],
        },
        "summary": {
            "total_rows": len(approved),
            "changed_by_human_approval_rows": len(changed),
            "human_review_basis_counts": dict(status_counts),
            "structure_counts": dict(structure_counts),
            "bad_structures": bad_structures,
            "duplicate_chars": duplicate_chars[:20],
            "duplicate_unicode": duplicate_unicode[:20],
            "source_table_rows_written": 0,
            "cnbe32_rows_written": 0,
            "database_rebuild_allowed": False,
        },
        "checks": checks,
        "outputs": {
            "approved_packet_csv": str(APPROVED_PACKET),
            "approved_changes_csv": str(APPROVED_CHANGES),
            "approved_workbook_xlsx": str(APPROVED_XLSX),
            "json_report": str(JSON_REPORT),
            "markdown_report": str(MD_REPORT),
        },
        "decision": {
            "may_use_as_8105_agent_structure_candidate_baseline": all(checks.values()),
            "may_start_next_merge_plan": all(checks.values()),
            "may_write_cnbe32": False,
            "may_rebuild_database": False,
            "recommended_next_step": (
                "Prepare a separate source-merge plan that maps the approved "
                "Agent structure candidates into the repository structure model. "
                "Do not modify source tables or rebuild databases without a new "
                "explicit authorization."
            ),
        },
    }
    return approved, changed, report


def write_workbook(path: Path, approved: list[dict[str, Any]], changed: list[dict[str, Any]], report: dict[str, Any]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    rows = [
        ("overall_status", report["overall_status"]),
        ("decision", report["human_decision"]["decision_text"]),
        ("total_rows", report["summary"]["total_rows"]),
        ("changed_by_human_approval_rows", report["summary"]["changed_by_human_approval_rows"]),
        ("source_table_rows_written", 0),
        ("cnbe32_rows_written", 0),
        ("database_rebuild_allowed", False),
    ]
    ws.append(("field", "value"))
    for row in rows:
        ws.append(row)

    for title, rows_to_write in (("Approved Packet", approved), ("Changed Rows", changed)):
        sheet = wb.create_sheet(title)
        fieldnames: list[str] = []
        for row in rows_to_write:
            for key in row:
                if key not in fieldnames:
                    fieldnames.append(key)
        sheet.append(fieldnames)
        for row in rows_to_write:
            sheet.append([row.get(key, "") for key in fieldnames])

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
        for idx, column_cells in enumerate(sheet.columns, 1):
            values = [str(cell.value or "") for cell in column_cells[:80]]
            width = min(max([len(v) for v in values] + [8]) + 2, 42)
            sheet.column_dimensions[get_column_letter(idx)].width = width
    wb.save(path)


def render_markdown(report: dict[str, Any]) -> str:
    lines = [
        "# Hanzi Decomp v0.3 Human-Approved 8105 Agent Structure Packet",
        "",
        f"- Overall status: `{report['overall_status']}`",
        f"- Human decision: {report['human_decision']['decision_text']}",
        f"- Total rows: {report['summary']['total_rows']}",
        f"- Changed by human approval: {report['summary']['changed_by_human_approval_rows']}",
        f"- Source table rows written: {report['summary']['source_table_rows_written']}",
        f"- CNBE32 rows written: {report['summary']['cnbe32_rows_written']}",
        f"- Database rebuild allowed: `{report['summary']['database_rebuild_allowed']}`",
        "",
        "This artifact is an Agent structure candidate baseline for the 8105",
        "scope. It is not a national-standard source table, not a CNBE32 write,",
        "and not a SQLite rebuild.",
        "",
        "## Human Review Basis Counts",
        "",
        "| Basis | Rows |",
        "|---|---:|",
    ]
    for status, count in sorted(report["summary"]["human_review_basis_counts"].items()):
        lines.append(f"| `{status}` | {count} |")
    lines.extend(
        [
            "",
            "## Checks",
            "",
            "| Check | Result |",
            "|---|---:|",
        ]
    )
    for check, result in sorted(report["checks"].items()):
        lines.append(f"| `{check}` | `{result}` |")
    lines.extend(["", "## Decision", "", report["decision"]["recommended_next_step"], ""])
    return "\n".join(lines)


def run() -> dict[str, Any]:
    approved, changed, report = build()
    write_csv(APPROVED_PACKET, approved)
    write_csv(APPROVED_CHANGES, changed)
    write_json(JSON_REPORT, report)
    write_text(MD_REPORT, render_markdown(report))
    write_workbook(APPROVED_XLSX, approved, changed, report)
    return report


def main() -> None:
    report = run()
    print(report["overall_status"])
    print(f"total_rows={report['summary']['total_rows']}")
    print(f"changed_by_human_approval_rows={report['summary']['changed_by_human_approval_rows']}")
    print(f"cnbe32_rows_written={report['summary']['cnbe32_rows_written']}")
    print(f"database_rebuild_allowed={report['summary']['database_rebuild_allowed']}")


if __name__ == "__main__":
    main()
