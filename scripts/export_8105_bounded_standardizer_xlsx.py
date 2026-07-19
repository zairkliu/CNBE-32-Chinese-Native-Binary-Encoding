#!/usr/bin/env python3
"""Export the bounded 8105 standardizer review packet to XLSX.

The canonical review packet remains CSV so the workflow does not require Excel
libraries. This optional exporter is for human review convenience.
"""

from __future__ import annotations

import csv
import sys
from pathlib import Path

SOURCE_CSV = Path("review_packets/300_character_8105_pilot/8105_core_bounded_standardizer_review_packet.csv")
OUTPUT_XLSX = Path("review_packets/300_character_8105_pilot/8105_core_bounded_standardizer_review_packet.xlsx")
RECHECK_CSV = Path("review_packets/300_character_8105_pilot/8105_no_legacy_human_recheck_packet.csv")
RECHECK_XLSX = Path("review_packets/300_character_8105_pilot/8105_no_legacy_human_recheck_packet.xlsx")


def export_csv_to_xlsx(source_csv: Path, output_xlsx: Path, sheet_name: str) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError:
        print("openpyxl is not installed; CSV review packet remains canonical", file=sys.stderr)
        raise SystemExit(2)

    with source_csv.open(encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle))
    if not rows:
        raise SystemExit(f"no rows found in {source_csv}")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

    headers = list(rows[0])
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in rows:
        sheet.append([row.get(header, "") for header in headers])

    freeze_cell = "A2"
    sheet.freeze_panes = freeze_cell
    sheet.auto_filter.ref = sheet.dimensions

    for index, header in enumerate(headers, start=1):
        width = min(max(len(header) + 2, 12), 42)
        if header in {"kangxi_excerpt", "zhonghua_excerpt", "blocked_items"}:
            width = 48
        if header in {"character", "standard_rank", "gf0017_existing_score"}:
            width = 12
        sheet.column_dimensions[get_column_letter(index)].width = width

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_xlsx)
    print(f"wrote {output_xlsx}")


def main() -> None:
    export_csv_to_xlsx(SOURCE_CSV, OUTPUT_XLSX, "8105 pilot review")
    if RECHECK_CSV.exists():
        export_csv_to_xlsx(RECHECK_CSV, RECHECK_XLSX, "8105 recheck")


if __name__ == "__main__":
    main()
