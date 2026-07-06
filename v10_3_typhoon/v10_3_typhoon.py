#!/usr/bin/env python3
"""
CNBE-32 v10.3 台风"巴威"实时路径预测实验
纯 Python 实现（无 numpy/PyTorch 依赖）
"""
import math, json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = os.path.dirname(os.path.abspath(__file__))

# ============================================================
# 数据：台风巴威 (202609) 历史轨迹 + 三种预报场景
# ============================================================
HIST = [
    # (lat, lon, wind_mps, pressure_hPa, label)
    (15.0, 152.0, 30, 980, "强热带风暴"),
    (15.0, 151.8, 60, 920, "超强台风"),
    (14.5, 150.5, 62, 915, "第一次巅峰"),
    (15.0, 149.0, 60, 920, "维持"),
    (13.5, 147.7, 62, 915, "强度回升"),
    (14.1, 145.7, 62, 915, "当前"),
]

FORECASTS = {
    "A_直扑浙江": [
        (16.5, 140.0, 60, 915),
        (18.5, 134.0, 58, 918),
        (21.0, 128.0, 55, 922),
        (23.5, 122.5, 50, 928),
        (26.0, 120.0, 45, 935),
    ],
    "B_先台后闽": [
        (16.5, 140.0, 60, 915),
        (18.5, 134.0, 58, 918),
        (20.5, 129.0, 55, 922),
        (22.5, 124.5, 48, 930),
        (24.5, 121.0, 42, 938),
    ],
    "C_穿越宫古岛": [
        (16.5, 140.0, 60, 915),
        (18.5, 134.0, 58, 918),
        (20.8, 127.5, 55, 922),
        (23.0, 123.5, 50, 928),
        (25.5, 120.5, 45, 935),
    ],
}

# ============================================================
# 编码器
# ============================================================
BOUNDS = {"lat":(0,50),"lon":(100,180),"wind":(0,85),"press":(900,1020)}

def quantize(v, lo, hi, bits=8):
    m = (1<<bits)-1
    return int(round(max(0, min(1, (v-lo)/(hi-lo))) * m))

def cnbe_encode(lat, lon, wind, press):
    """4×8bit = 32bit CNBE 编码"""
    return (quantize(lat,0,50)<<24)|(quantize(lon,100,180)<<16)|(quantize(wind,0,85)<<8)|quantize(press,900,1020)

def raw_norm(lat, lon, wind, press):
    """归一化的原始值"""
    return [quantize(lat,0,50,8)/255, quantize(lon,100,180,8)/255,
            quantize(wind,0,85,8)/255, quantize(press,900,1020,8)/255]

# ============================================================
# 度量工具
# ============================================================
def haversine_km(lat1, lon1, lat2, lon2):
    R = 6371
    dlat = math.radians(lat2-lat1)
    dlon = math.radians(lon2-lon1)
    a = math.sin(dlat/2)**2 + math.cos(math.radians(lat1))*math.cos(math.radians(lat2))*math.sin(dlon/2)**2
    return 2 * R * math.atan2(math.sqrt(a), math.sqrt(1-a))

def hamming_dist(a, b):
    return bin(a ^ b).count('1')

def euclid_dist(a, b):
    return math.sqrt(sum((x-y)**2 for x,y in zip(a,b)))

# ============================================================
# 预测器
# ============================================================
def predict(hist, forecast, encode_func, dist_func):
    """
    对每个预报点，找最近的历史状态，用其后续移动向量做预测
    返回: [(预测lat, 预测lon, 真实lat, 真实lon, 误差km), ...]
    """
    # 历史状态编码
    h_codes = [encode_func(lat, lon, w, p) for lat,lon,w,p,_ in HIST]
    # 历史移动向量
    h_moves = [(HIST[i+1][0]-HIST[i][0], HIST[i+1][1]-HIST[i][1]) for i in range(len(HIST)-1)]
    h_move_codes = [(h_codes[i], h_codes[i+1], h_moves[i]) for i in range(len(h_moves))]
    
    results = []
    for f_lat, f_lon, f_w, f_p in forecast:
        f_code = encode_func(f_lat, f_lon, f_w, f_p)
        
        # 找最近的历史状态
        best_idx = min(range(len(h_codes)-1), key=lambda i: dist_func(h_codes[i], f_code))
        
        # 用最佳匹配的后续移动做预测
        dlat, dlon = h_moves[best_idx]
        p_lat = f_lat + dlat
        p_lon = f_lon + dlon
        err = haversine_km(p_lat, p_lon, f_lat, f_lon)
        results.append((p_lat, p_lon, f_lat, f_lon, err))
    
    return results

# ============================================================
# 实验主流程
# ============================================================
def run_all():
    print("=" * 60)
    print("CNBE-32 v10.3 台风巴威路径预测实验")
    print("=" * 60)
    print(f"\n历史轨迹: {len(HIST)} 个时间点")
    
    all_results = {}
    
    for name, fcast in FORECASTS.items():
        print(f"\n{'─'*60}")
        print(f"场景: {name}")
        
        # 四种编码方案
        for method, encode, dist, label in [
            ("CNBE", cnbe_encode, hamming_dist, "CNBE-32 编码"),
            ("RAW", raw_norm, euclid_dist, "原始数值"),
        ]:
            results = predict(HIST, fcast, encode, dist)
            avg_err = sum(r[4] for r in results) / len(results)
            print(f"  {label:15s}: 平均误差 {avg_err:.0f} km")
            
            if name not in all_results:
                all_results[name] = {}
            all_results[name][method] = {
                "avg_error_km": round(avg_err, 1),
                "details": [round(r[4], 1) for r in results]
            }
    
    # 汇总
    print(f"\n{'='*60}")
    print("汇总")
    print(f"{'='*60}")
    print(f"\n{'场景':15s} {'CNBE(km)':12s} {'RAW(km)':12s} {'差异':12s}")
    print(f"{'─'*55}")
    for name in FORECASTS:
        c = all_results[name]["CNBE"]["avg_error_km"]
        r = all_results[name]["RAW"]["avg_error_km"]
        diff = c - r
        winner = "CNBE✓" if diff < 0 else "RAW "
        print(f"{name:15s} {c:>8.0f} km  {r:>8.0f} km  {diff:>+7.0f} km  [{winner}]")
    
    # 生成 XLSX 报告
    wb = Workbook()
    hf = Font(bold=True, size=11, color="FFFFFF")
    hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    cf = Font(size=10)
    ca = Alignment(horizontal="center")
    tb = Border(left=Side(style="thin"),right=Side(style="thin"),
                 top=Side(style="thin"),bottom=Side(style="thin"))
    
    ws = wb.active
    ws.title = "台风预测实验"
    headers = ["场景","方法","平均误差(km)","误差明细"]
    for i,h in enumerate(headers,1):
        c=ws.cell(row=1,column=i,value=h); c.font=hf; c.fill=hfill; c.alignment=ca; c.border=tb
    
    row = 2
    for name in FORECASTS:
        for method in ["CNBE","RAW"]:
            d = all_results[name][method]
            for j,v in enumerate([name, method, d["avg_error_km"], str(d["details"])],1):
                c=ws.cell(row=row,column=j,value=v); c.font=cf; c.alignment=ca; c.border=tb
            row += 1
    
    ws2 = wb.create_sheet("对比")
    for i,h in enumerate(["场景","CNBE(km)","RAW(km)","差异(km)","优胜"],1):
        c=ws2.cell(row=1,column=i,value=h); c.font=hf; c.fill=hfill; c.alignment=ca; c.border=tb
    for i,name in enumerate(FORECASTS,2):
        c=all_results[name]["CNBE"]["avg_error_km"]
        r=all_results[name]["RAW"]["avg_error_km"]
        for j,v in enumerate([name,c,r,round(c-r,1),"CNBE✓" if c<r else "RAW"],1):
            c2=ws2.cell(row=i,column=j,value=v); c2.font=cf; c2.alignment=ca; c2.border=tb
    
    xlsx_path = os.path.join(OUT, "v10_3_typhoon_report.xlsx")
    wb.save(xlsx_path)
    print(f"\n报告: {xlsx_path}")
    
    # JSON report
    with open(os.path.join(OUT, "v10_3_report.json"), "w") as f:
        json.dump({"experiment": "v10.3 台风巴威路径预测", "results": all_results}, f, ensure_ascii=False, indent=2)
    print(f"报告: {os.path.join(OUT, 'v10_3_report.json')}")
    
    return all_results

if __name__ == "__main__":
    run_all()
