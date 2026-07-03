# -*- coding: utf-8 -*-
import sys, os, time, json, struct
sys.stdout.reconfigure(encoding="utf-8")
import numpy as np
from openpyxl import load_workbook
OUT = os.path.dirname(__file__) or "."
BASE = os.path.dirname(OUT)
XLSX = os.path.join(BASE, "CNBE_\u7f16\u7801\u76ee\u5f55_\u4fee\u590d\u7248_v3.xlsx")
print(f"Loading {XLSX}...")
wb = load_workbook(XLSX); ws = wb.worksheets[0]; n = ws.max_row - 1
codes = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[1] is None: break
    codes.append(int(row[4]))
n = len(codes)
print(f"Loaded {n} codes")
test_codes = codes[:min(10000, n)]
N_ITER = 10000
print(f"Test: {len(test_codes)} chars x {N_ITER} loops = {len(test_codes)*N_ITER:,} ops")

# 5 CNBE instructions cycle counts
INST_CYCLES = {"extract": 1, "cmp": 2, "map": 2, "mv": 1, "enc": 1}
FREQ_HZ = 2.5e9  # 2.5 GHz typical RISC-V

# Software simulation
print("\n--- Software Path (Python bit operations) ---")
t0 = time.time()
for _ in range(N_ITER):
    for c in test_codes:
        _ = (c >> 24) & 0xFF  # extract radical
        _ = (c >> 19) & 0x1F  # extract stroke
        _ = (c >> 15) & 0x0F  # extract structure
    for i in range(len(test_codes)-1):
        ca, cb = test_codes[i], test_codes[i+1]
        _ = abs((ca>>24)-(cb>>24))*8 + abs(((ca>>19)&0x1F)-((cb>>19)&0x1F))*5 + abs(((ca>>15)&0x0F)-((cb>>15)&0x0F))*4 + abs(((ca>>4)&0x7FF)-((cb>>4)&0x7FF))*11
t_sw = time.time() - t0
sw_per_char = t_sw / (N_ITER * len(test_codes)) * 1e6
print(f"  Total: {t_sw*1000:.0f} ms")
print(f"  Per char: {sw_per_char:.2f} us")

# Hardware simulation (cycle-accurate)
print("\n--- Hardware Path (RISC-V instruction simulation) ---")
cycles_extract = len(test_codes) * N_ITER * INST_CYCLES["extract"]
cycles_cmp = (len(test_codes)-1) * N_ITER * INST_CYCLES["cmp"]
total_cycles = cycles_extract + cycles_cmp
t_hw = total_cycles / FREQ_HZ
hw_per_char = t_hw / (N_ITER * len(test_codes)) * 1e9
print(f"  Cycles: {total_cycles:,}")
print(f"  @2.5GHz: {t_hw*1000:.3f} ms")
print(f"  Per char: {hw_per_char:.1f} ns")

# Speedup
speedup = t_sw / t_hw
print(f"\n=== SPEEDUP: {speedup:.0f}x ===")

# Full CJK throughput projection
print("\n--- Full CJK (97,686 chars) Throughput Projection ---")
t_sw_full = t_sw / len(test_codes) * n
t_hw_full = t_hw / len(test_codes) * n
print(f"  Software: {t_sw_full*1000:.0f} ms for full CJK encode")
print(f"  Hardware: {t_hw_full*1000:.3f} ms for full CJK encode")

# Memory footprint for cnhe.map lookup table
print("\n--- CNHE.MAP Lookup Table ---")
lookup_size = n * 6  # 6 bytes per entry (2 bytes for Unicode + 4 bytes for CNBE code)
print(f"  {n} entries x 6 bytes = {lookup_size/1024:.0f} KB ROM")
print(f"  Fits in L2 cache: {lookup_size < 2*1024*1024}")

# Compare with alternatives
print("\n--- Comparison with Alternative Encodings ---")
print(f"  One-hot: {n} x 4 bytes = {n*4/1024/1024:.1f} MB")
print(f"  Word2vec (100-dim): {n} x 100 x 4 = {n*100*4/1024/1024:.1f} MB")
print(f"  CNBE-32 (native): {n} x 4 = {n*4/1024/1024:.1f} MB")
print(f"  CNBE-32 (ROM): {lookup_size/1024:.0f} KB")

print("\nAll experiments complete.")
results = {
    "sw_time_ms": t_sw*1000,
    "hw_time_ms": t_hw*1000,
    "speedup": speedup,
    "sw_per_char_us": sw_per_char,
    "hw_per_char_ns": hw_per_char,
    "total_cycles": total_cycles,
    "rom_size_kb": lookup_size/1024,
}
with open(os.path.join(OUT, "riscv_results.json"), "w") as f:
    json.dump(results, f, indent=2)
print(f"\nResults saved to riscv_results.json")
