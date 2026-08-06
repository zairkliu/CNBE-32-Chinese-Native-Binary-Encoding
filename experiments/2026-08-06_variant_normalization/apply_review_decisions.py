#!/usr/bin/env python3
"""Apply the owner's human review decisions to the coverage-gap workbook and packet.

Review content (2026-08-06) is embedded below. The script:
  1. updates the `待人工复核` sheet in the xlsx;
  2. writes a `review` record and updated `proposed` fields into
     coverage_remediation_packet.json;
  3. prints the resolved radix/struct codes for verification.
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

EXP = Path(__file__).resolve().parent
OUT_XLSX = EXP.parents[1] / "evidence" / "validation" / "CNBE覆盖缺口人工复核_2026-08-06.xlsx"
PACKET = EXP / "coverage_remediation_packet.json"

STRUCT_ORDER = [
    "独体字", "上下", "上中下", "左右", "左中右", "左上包", "右上包",
    "左三包", "左下包", "上三包", "下三包", "全包围", "镶嵌",
]
STRUCT_CODE = {name: i for i, name in enumerate(STRUCT_ORDER)}

# owner review: char, radix_name, strokes, structure
REVIEW = [
    ("也", "乙", 3, "独体"),
    ("爲", "爫", 12, "独体"),
    ("東", "一", 8, "独体"),
    ("飛", "飞", 9, "独体"),
    ("巳", "巳", 3, "独体"),
    ("巴", "巳", 4, "独体"),
    ("㻛", "王", 12, "左右"),
    ("民", "氏", 5, "独体"),
    ("車", "車", 7, "独体"),
    ("㳅", "氵", 7, "左右"),
    ("㸔", "爪", 11, "左上包"),
    ("䄂", "礻", 10, "左右"),
    ("龜", "龜", 17, "独体"),
    ("㮣", "木", 13, "上下"),
    ("㴠", "氵", 12, "左右"),
    ("肅", "肀", 13, "独体"),
]


def norm_struct(s: str) -> str:
    return "独体字" if s == "独体" else s


def load_radix_map(path: Path) -> dict[str, int]:
    data = json.loads(path.read_text(encoding="utf-8"))
    result: dict[str, int] = {}
    for rec in data.get("records", []):
        if rec.get("code") is not None:
            result[rec["radical"]] = rec["code"]
            if rec.get("canonical_radical"):
                result.setdefault(rec["canonical_radical"], rec["code"])
    return result


def load_unihan_krs(path: Path) -> dict[str, int]:
    result: dict[str, int] = {}
    for line in path.read_text(encoding="utf-8", errors="replace").splitlines():
        if line.startswith("#"):
            continue
        parts = line.split("\t")
        if len(parts) < 3 or parts[1] != "kRSUnicode":
            continue
        m = re.match(r"\d+", parts[2])
        if m:
            result[parts[0]] = int(m.group(0))
    return result


def main() -> None:
    radix_map = load_radix_map(EXP.parents[1] / "evidence" / "8105" / "cnbe8105_radical_code_map.json")
    build = EXP.parent / "2026-08-05_scheme_comparison" / "build"
    unihan_krs = load_unihan_krs(build / "Unihan_IRGSources.txt")

    decisions = {}
    for c, radix_name, strokes, structure in REVIEW:
        struct_name = norm_struct(structure)
        ucp = f"U+{ord(c):04X}"
        radix_code = radix_map.get(radix_name)
        source = "radix_map"
        if radix_code is None and ucp in unihan_krs:
            radix_code = unihan_krs[ucp]
            source = "unihan_fallback"
        struct_code = STRUCT_CODE.get(struct_name)
        proposed = {
            "radix": radix_code,
            "radix_name": radix_name,
            "strokes": strokes,
            "struct_name": struct_name,
            "struct_type": struct_code,
            "index": (ord(c) - 0x4E00) % 2048,
            "ext": 0,
            "track": "provisional",
        }
        decisions[c] = {
            "decision": "批准",
            "reviewer": "zairkliu",
            "reviewed_at": "2026-08-06",
            "reviewed_radix_name": radix_name,
            "reviewed_strokes": strokes,
            "reviewed_structure": struct_name,
            "radix_code": radix_code,
            "radix_source": source,
            "struct_code": struct_code,
            "proposed": proposed,
        }
        print(c, radix_name, strokes, struct_name, "-> radix", radix_code, source, "struct", struct_code)

    packet = json.loads(PACKET.read_text(encoding="utf-8"))
    updated = 0
    for entry in packet["entries"]:
        if entry["char"] in decisions:
            entry["review_status"] = "APPROVED"
            entry["review"] = decisions[entry["char"]]
            entry["proposed"] = decisions[entry["char"]]["proposed"]
            updated += 1
    packet["summary"]["reviewed_approved"] = updated
    PACKET.write_text(json.dumps(packet, ensure_ascii=False, indent=2), encoding="utf-8")

    wb = load_workbook(OUT_XLSX)
    ws = wb["待人工复核"]
    filled = 0
    for row in ws.iter_rows(min_row=2):
        c = row[0].value
        if c not in decisions:
            continue
        d = decisions[c]
        row[22].value = "批准"
        row[15].value = d["radix_code"]
        row[16].value = d["reviewed_radix_name"]
        row[17].value = d["reviewed_strokes"]
        row[18].value = d["reviewed_structure"]
        row[19].value = d["struct_code"]
        row[20].value = d["proposed"]["index"]
        row[21].value = "provisional"
        row[23].value = d["reviewer"]
        row[24].value = "人工复核录入 2026-08-06"
        filled += 1
    wb.save(OUT_XLSX)
    print("packet_updated", updated, "xlsx_filled", filled)


if __name__ == "__main__":
    main()
