#!/usr/bin/env python3
"""Export the 8105 remaining legacy-track rows to a human review workbook.

Reads the read-only completion packet produced by
scripts/advance_8105_remaining.py and emits an editable XLSX with current DB
fields, standard evidence, Unihan cross-references, and blank proposal fields
for the reviewer. No release database write is performed here.
"""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

EXP = Path(__file__).resolve().parent
REPO = EXP.parents[1]
PACKET = REPO / "evidence" / "8105" / "8105_REMAINING_503_COMPLETION_PACKET.json"
OUT = REPO / "evidence" / "validation" / "CNBE8105_LEGACY_REVIEW_2026-08-06.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="2F5597")
REVIEW_FILL = PatternFill("solid", fgColor="FCE4D6")
HEADER_FONT = Font(color="FFFFFF", bold=True)

HEADERS = [
    "汉字",
    "Unicode",
    "8105序号",
    "当前CNBE",
    "当前部首码",
    "当前部首名",
    "当前笔画",
    "当前结构",
    "当前轨道",
    "标准部首",
    "标准笔画",
    "标准结构",
    "标准分解",
    "证据状态",
    "问题列表",
    "Unihan部首",
    "Unihan部首码",
    "Unihan部首名",
    "Unihan笔画",
    "CJK分解",
    "建议部首码",
    "建议部首名",
    "建议笔画",
    "建议结构",
    "建议结构码",
    "建议idx",
    "建议轨道",
    "复核决定(批准/驳回/修改)",
    "复核人",
    "备注",
]

WIDTHS = [
    8, 10, 10, 12, 10, 10, 10, 12, 10, 10, 10, 12, 20, 14, 24, 14, 12, 12, 12, 22,
    12, 12, 10, 12, 12, 10, 12, 22, 10, 24,
]


def row_from_entry(e: dict) -> list:
    cur = e.get("current") or {}
    std = e.get("standard_evidence") or {}
    cross = e.get("cross_reference") or {}
    return [
        e["char"],
        e["unicode"],
        e.get("standard_rank") or "",
        cur.get("cnbe", ""),
        cur.get("radix", ""),
        cur.get("radix_name", ""),
        cur.get("strokes", ""),
        cur.get("struct_name", ""),
        cur.get("track", ""),
        std.get("radical", ""),
        std.get("stroke_count", ""),
        std.get("structure", ""),
        std.get("decomposition", ""),
        std.get("evidence_status", ""),
        "\n".join(std.get("issues", [])),
        cross.get("kRSUnicode", ""),
        cross.get("krs_radical_code", ""),
        cross.get("krs_radical_name", ""),
        cross.get("kTotalStrokes", ""),
        cross.get("cjk_decomp", ""),
        "", "", "", "", "", "", "provisional", "", "", "",
    ]


def write_sheet(ws, rows: list[dict], title: str) -> None:
    ws.title = title
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for e in rows:
        ws.append(row_from_entry(e))
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.fill = REVIEW_FILL
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def main() -> None:
    packet = json.loads(PACKET.read_text(encoding="utf-8"))
    entries = packet["entries"]

    wb = Workbook()
    ws_info = wb.active
    ws_info.title = "授权与说明"
    info = [
        ["CNBE 8105 legacy 剩余行人工复核授权"],
        ["授权人", "项目负责人"],
        ["授权日期", "2026-08-06"],
        ["范围", "8105 规范字表内仍为 legacy 轨的 491 行"],
        ["数据来源", "8105_REMAINING_503_COMPLETION_PACKET.json（只读生成，未写发布库）"],
        ["门禁", "本次授权仅允许形成复核结论与候选字段；发布库写入需另走治理流程"],
        ["证据等级", "standard_evidence + Unihan kRSUnicode/kTotalStrokes + CHISE cjk_decomp 交叉参考"],
        ["填写要求", "复核决定列填写：批准 / 驳回 / 修改；修改时在建议字段给出新值"],
        ["输出", "复核完成后返回本工作簿，由项目方生成候选库副本并验证"],
    ]
    for row in info:
        ws_info.append(row)
    ws_info.column_dimensions["A"].width = 18
    ws_info.column_dimensions["B"].width = 80
    for cell in ws_info[1]:
        cell.font = Font(bold=True, size=14)

    write_sheet(wb.create_sheet(), entries, "待人工审核491")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    auth = {
        "authorizer": "项目负责人",
        "authorized_at": datetime.now().isoformat(timespec="seconds"),
        "scope": "8105 remaining legacy-track rows",
        "legacy_rows": len(entries),
        "gate": "review_conclusion_only_no_release_db_write",
        "workbook": str(OUT),
    }
    (EXP / "review_authorization_legacy.json").write_text(
        json.dumps(auth, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(json.dumps(auth, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
