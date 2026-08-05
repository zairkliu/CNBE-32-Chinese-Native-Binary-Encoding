#!/usr/bin/env python3
"""Export the CNBE coverage-gap remediation packet to a human-review workbook.

Workbook sheets:
  1. 授权与说明 - owner authorization and review rules
  2. 全部812     - every gap char with current/proposed/evidence fields
  3. 待人工复核   - UPGRADE_REVIEW (10) + INSERT_CANDIDATE (6)
  4. 候选批量批准 - UPGRADE_CANDIDATE (796)
"""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

EXP = Path(__file__).resolve().parent
OUT = EXP.parents[1] / "evidence" / "validation" / "CNBE覆盖缺口人工复核_2026-08-06.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="2F5597")
REVIEW_FILL = PatternFill("solid", fgColor="FCE4D6")
INSERT_FILL = PatternFill("solid", fgColor="FFF2CC")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def row_from_entry(e: dict) -> list:
    cur = e.get("current") or {}
    ev = e.get("evidence") or {}
    prop = e.get("proposed") or {}
    return [
        e["char"],
        e["codepoint"],
        e["count"],
        "是" if e["in_db"] else "否",
        e["db_track"] or "",
        cur.get("radix_name", ""),
        cur.get("strokes", ""),
        cur.get("struct_name", ""),
        ev.get("kRSUnicode", ""),
        ev.get("krs_radical_code", ""),
        ev.get("kTotalStrokes", ""),
        ev.get("ids", ""),
        ev.get("inferred_structure", ""),
        e["action"],
        e["evidence_grade"],
        prop.get("radix", ""),
        prop.get("radix_name", ""),
        prop.get("strokes", ""),
        prop.get("struct_name", ""),
        prop.get("struct_type", ""),
        prop.get("index", ""),
        prop.get("track", ""),
        "",
        "",
        "",
    ]


HEADERS = [
    "汉字",
    "Unicode",
    "出现次数",
    "当前在库",
    "当前轨道",
    "当前部首",
    "当前笔画",
    "当前结构",
    "证据kRSUnicode",
    "证据部首码",
    "证据总笔画",
    "IDS",
    "推断结构",
    "动作",
    "证据等级",
    "建议部首码",
    "建议部首名",
    "建议笔画",
    "建议结构",
    "建议结构码",
    "建议idx",
    "建议轨道",
    "复核决定(批准/驳回/修改)",
    "复核人",
    "备注",
]


def write_sheet(ws, rows: list[dict], title: str) -> None:
    ws.title = title
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for e in rows:
        ws.append(row_from_entry(e))
    for row in ws.iter_rows(min_row=2):
        action = row[13].value
        if action == "UPGRADE_REVIEW":
            for cell in row:
                cell.fill = REVIEW_FILL
        elif action == "INSERT_CANDIDATE":
            for cell in row:
                cell.fill = INSERT_FILL
    widths = [8, 10, 10, 10, 12, 10, 10, 12, 14, 12, 12, 16, 12, 16, 18, 12, 12, 10, 12, 12, 10, 12, 18, 10, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def main() -> None:
    packet = json.loads((EXP / "coverage_remediation_packet.json").read_text(encoding="utf-8"))
    entries = packet["entries"]
    review = [e for e in entries if e["action"] in ("UPGRADE_REVIEW", "INSERT_CANDIDATE")]
    batch = [e for e in entries if e["action"] == "UPGRADE_CANDIDATE"]

    wb = Workbook()
    ws_info = wb.active
    ws_info.title = "授权与说明"
    info = [
        ["CNBE 覆盖缺口人工复核授权"],
        ["授权人", "项目负责人"],
        ["授权日期", "2026-08-06"],
        ["范围", "37 页永乐大典真值中不在 CNBE 标准轨的 812 个独有字"],
        ["复核对象", "UPGRADE_REVIEW 10 个 + INSERT_CANDIDATE 6 个 + UPGRADE_CANDIDATE 796 个"],
        ["门禁", "本次授权只允许形成复核结论与候选数据；发布库写入需另走治理流程"],
        ["证据等级", "cross_reference_unihan_ids（Unihan + CHISE IDS 交叉参考，非国家标准）"],
        ["填写要求", "复核决定列填写：批准 / 驳回 / 修改；修改时在备注给出建议字段"],
        ["输出", "复核完成后返回本工作簿，由项目方生成候选库副本并验证"],
    ]
    for row in info:
        ws_info.append(row)
    ws_info.column_dimensions["A"].width = 18
    ws_info.column_dimensions["B"].width = 80
    for cell in ws_info[1]:
        cell.font = Font(bold=True, size=14)

    write_sheet(wb.create_sheet(), entries, "全部812")
    write_sheet(wb.create_sheet(), review, "待人工复核")
    write_sheet(wb.create_sheet(), batch, "候选批量批准")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    auth = {
        "authorizer": "项目负责人",
        "authorized_at": datetime.now().isoformat(timespec="seconds"),
        "scope": "CNBE coverage-gap human review for 812 unique truth chars",
        "gate": "review_conclusion_only_no_release_db_write",
        "review_required": len(review),
        "batch_candidates": len(batch),
        "workbook": str(OUT),
    }
    (EXP / "review_authorization.json").write_text(
        json.dumps(auth, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(json.dumps(auth, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
