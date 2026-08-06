#!/usr/bin/env python3
"""Import the second-round fallback-radical review (66 rows) into the packet.

All 66 rows were approved with a revised radical name. The script updates the
packet, rebuilds the candidate DB copy, re-verifies roundtrip, and regenerates
the main review workbook.
"""

from __future__ import annotations

import json
import os
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, str(Path(__file__).resolve().parent))

import apply_batch_review as abr  # noqa: E402
import build_candidate_db as bcd  # noqa: E402
import export_review_xlsx as erx  # noqa: E402

EXP = Path(__file__).resolve().parent
PACKET = EXP / "coverage_remediation_packet.json"


def find_fallback_xlsx() -> Path:
    one = Path("C:/Users/zairk/OneDrive")
    desk = None
    for name in os.listdir(one):
        if len(name) >= 2 and name[0] == "\u684c" and name[1] == "\u9762":
            desk = one / name
            break
    for f in os.listdir(desk):
        if f.startswith("CNBE") and "\u90e8\u9996\u540d" in f and f.endswith(".xlsx"):
            return desk / f
    raise FileNotFoundError("fallback review xlsx not found")


def main() -> None:
    xlsx_path = find_fallback_xlsx()
    radix_map = abr.load_radix_map(EXP.parents[1] / "evidence" / "8105" / "cnbe8105_radical_code_map.json")
    krs = abr.load_unihan_krs(
        EXP.parent / "2026-08-05_scheme_comparison" / "build" / "Unihan_IRGSources.txt"
    )

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[1]]
    updates = {}
    for row in ws.iter_rows(min_row=2):
        c = row[0].value
        if c is None:
            continue
        decision = row[11].value
        name = row[12].value
        strokes = row[13].value
        structure = row[14].value
        note = row[15].value or ""
        updates[c] = {
            "decision": str(decision).strip() if decision else "批准",
            "name": name,
            "strokes": strokes,
            "structure": structure,
            "note": note,
        }

    packet = json.loads(PACKET.read_text(encoding="utf-8"))
    applied = 0
    for entry in packet["entries"]:
        if entry["char"] not in updates:
            continue
        u = updates[entry["char"]]
        review = entry.setdefault("review", {})
        name = u["name"] or review.get("reviewed_radix_name")
        strokes = u["strokes"] if u["strokes"] is not None else review.get("reviewed_strokes")
        structure = u["structure"] or review.get("reviewed_structure")
        if structure == "独体":
            structure = "独体字"
        radix_code = radix_map.get(name) or abr.RADIX_SUPPLEMENT.get(name)
        source = "radix_map" if name in radix_map else "supplement"
        mismatch = False
        if radix_code is None:
            ucp = f"U+{ord(entry['char']):04X}"
            if ucp in krs:
                radix_code = krs[ucp]
                source = "unihan_fallback"
                mismatch = True
        struct_code = abr.STRUCT_CODE.get(structure)
        review["decision"] = u["decision"]
        review["reviewer"] = "zairkliu"
        review["reviewed_at"] = "2026-08-06"
        review["reviewed_radix_name"] = name
        review["reviewed_strokes"] = int(strokes)
        review["reviewed_structure"] = structure
        review["radix_code"] = radix_code
        review["radix_source"] = source
        review["radix_name_mismatch"] = mismatch
        review["struct_code"] = struct_code
        review["second_round"] = {
            "decision": u["decision"],
            "reviewer": "zairkliu",
            "reviewed_at": "2026-08-06",
            "note": u["note"],
        }
        entry["review_status"] = "APPROVED"
        entry["proposed"] = {
            "radix": radix_code,
            "radix_name": name,
            "strokes": int(strokes),
            "struct_name": structure,
            "struct_type": struct_code,
            "index": (ord(entry["char"]) - 0x4E00) % 2048,
            "ext": 0,
            "track": "provisional",
        }
        applied += 1
    packet["summary"]["fallback_review_approved"] = applied
    PACKET.write_text(json.dumps(packet, ensure_ascii=False, indent=2), encoding="utf-8")

    bcd.main()
    erx.main()
    summary = {
        "xlsx": str(xlsx_path),
        "rows": len(updates),
        "packet_updated": applied,
        "unresolved_after_review": sum(1 for e in packet["entries"] if e.get("review", {}).get("radix_name_mismatch")),
    }
    (EXP / "fallback_review_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
