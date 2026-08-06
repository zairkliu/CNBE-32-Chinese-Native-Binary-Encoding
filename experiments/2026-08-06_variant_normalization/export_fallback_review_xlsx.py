#!/usr/bin/env python3
"""Export a focused review workbook for the 66 Unihan-fallback radical names.

These rows passed batch review but their reviewed radical name is not in the
project radical map, so the radix code was resolved via Unihan kRSUnicode. The
workbook lets the reviewer confirm or replace the radical name/code.
"""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

EXP = Path(__file__).resolve().parent
OUT = EXP.parents[1] / "evidence" / "validation" / "CNBE部首名回退复核_2026-08-06.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="C55A11")
HEADER_FONT = Font(color="FFFFFF", bold=True)

FALLBACK_NAMES = {
    5: "乙",
    71: "无",
    75: "木",
    79: "殳",
    108: "皿",
    120: "糸",
    129: "聿",
    178: "韋",
}

HEADERS = [
    "汉字",
    "Unicode",
    "出现次数",
    "人工审核部首名",
    "回退部首码",
    "回退部首名",
    "建议笔画",
    "建议结构",
    "建议结构码",
    "建议idx",
    "动作",
    "复核决定(批准/修改)",
    "复核部首名",
    "复核笔画",
    "复核结构",
    "备注",
]


def main() -> None:
    packet = json.loads((EXP / "coverage_remediation_packet.json").read_text(encoding="utf-8"))
    flagged = [e for e in packet["entries"] if e.get("review", {}).get("radix_name_mismatch")]
    flagged.sort(key=lambda e: (-e["count"], e["char"]))

    wb = Workbook()
    ws_info = wb.active
    ws_info.title = "说明"
    info = [
        ["CNBE 部首名回退复核（66 条）"],
        ["用途", "批量审核中 66 个部首名不在项目映射表，部首码由 Unihan kRSUnicode 回退"],
        ["填写要求", "复核决定填 批准 / 修改；修改时在 复核部首名/笔画/结构 填写建议值"],
        ["输出", "复核完成后返回工作簿，由项目方更新候选库副本并重新验证"],
        ["注意", "回退部首名仅为参考，可能使用异体/部件形式"],
    ]
    for row in info:
        ws_info.append(row)
    ws_info.column_dimensions["A"].width = 18
    ws_info.column_dimensions["B"].width = 90
    for cell in ws_info[1]:
        cell.font = Font(bold=True, size=14)

    ws = wb.create_sheet("回退部首复核")
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for e in flagged:
        review = e["review"]
        prop = e.get("proposed") or {}
        code = review["radix_code"]
        ws.append(
            [
                e["char"],
                e["codepoint"],
                e["count"],
                review["reviewed_radix_name"],
                code,
                FALLBACK_NAMES.get(code, ""),
                review["reviewed_strokes"],
                review["reviewed_structure"],
                review["struct_code"],
                prop.get("index", ""),
                e["action"],
                "",
                "",
                "",
                "",
                "",
            ]
        )
    widths = [8, 10, 10, 14, 10, 12, 10, 12, 10, 10, 16, 18, 12, 10, 12, 24]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print("rows", len(flagged), "out", OUT)


if __name__ == "__main__":
    main()
