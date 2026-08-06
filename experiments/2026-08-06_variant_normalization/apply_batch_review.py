#!/usr/bin/env python3
"""Import the owner's batch human review (Desktop xlsx) into the coverage packet.

The batch sheet fills radical name / strokes / structure for 796 UPGRADE_CANDIDATE
rows. This script treats filled review fields as approval, resolves radix/struct
codes, and writes `review` records plus updated `proposed` fields into
coverage_remediation_packet.json.
"""

from __future__ import annotations

import json
import os
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

EXP = Path(__file__).resolve().parent
PACKET = EXP / "coverage_remediation_packet.json"
REPO = EXP.parents[1]

STRUCT_ORDER = [
    "独体字", "上下", "上中下", "左右", "左中右", "左上包", "右上包",
    "左三包", "左下包", "上三包", "下三包", "全包围", "镶嵌",
]
STRUCT_CODE = {name: i for i, name in enumerate(STRUCT_ORDER)}

# Kangxi 214 radical name -> code supplement for names absent from the project map.
RADIX_SUPPLEMENT = {
    "一": 1, "乙": 5, "乚": 5, "二": 7, "人": 9, "儿": 10, "入": 11, "八": 12,
    "刀": 18, "刂": 18, "力": 19, "口": 30, "囗": 31, "土": 32, "士": 33,
    "夂": 34, "夕": 36, "大": 37, "女": 38, "子": 39, "宀": 40, "寸": 41,
    "小": 42, "尢": 43, "尸": 44, "山": 46, "工": 48, "己": 49, "巳": 49,
    "巾": 50, "干": 51, "广": 53, "廴": 54, "廾": 55, "弋": 56, "弓": 57,
    "彡": 59, "彳": 60, "心": 61, "戈": 62, "戶": 63, "户": 63, "手": 64,
    "扌": 64, "支": 65, "攴": 66, "文": 67, "斗": 68, "斤": 69, "方": 70,
    "无": 71, "日": 72, "曰": 73, "月": 74, "木": 75, "欠": 76, "止": 77,
    "歹": 78, "殳": 79, "毋": 80, "比": 81, "毛": 82, "氏": 83, "气": 84,
    "水": 85, "氵": 85, "火": 86, "灬": 86, "爪": 87, "爫": 87, "父": 88,
    "爻": 89, "爿": 90, "片": 91, "牙": 92, "牛": 93, "犬": 94, "犭": 94,
    "玄": 95, "王": 96, "玉": 96, "瓦": 98, "甘": 99, "生": 100, "用": 101,
    "田": 102, "疋": 103, "疒": 104, "癶": 105, "白": 106, "皮": 107,
    "皿": 108, "目": 109, "矛": 110, "矢": 111, "石": 112, "示": 113,
    "礻": 113, "禸": 114, "禾": 115, "穴": 116, "立": 117, "竹": 118,
    "米": 119, "糸": 120, "纟": 120, "缶": 121, "网": 122, "羊": 123,
    "羽": 124, "老": 125, "而": 126, "耒": 127, "耳": 128, "聿": 129,
    "肀": 129, "肉": 130, "臣": 131, "自": 132, "至": 133, "臼": 134,
    "舌": 135, "舛": 136, "舟": 137, "艮": 138, "色": 139, "艸": 140,
    "艹": 140, "虍": 141, "虫": 142, "血": 143, "行": 144, "衣": 145,
    "衤": 145, "西": 146, "見": 147, "見": 147, "角": 148, "言": 149,
    "谷": 150, "豆": 151, "豕": 152, "豸": 153, "貝": 154, "赤": 155,
    "走": 156, "足": 157, "身": 158, "車": 159, "辛": 160, "辰": 161,
    "辵": 162, "辶": 162, "邑": 163, "酉": 164, "釆": 165, "里": 166,
    "金": 167, "長": 168, "門": 169, "阜": 170, "阝": 170, "隶": 171,
    "隹": 172, "雨": 173, "青": 174, "非": 175, "面": 176, "革": 177,
    "韋": 178, "韭": 179, "音": 180, "頁": 181, "風": 182, "飛": 183,
    "飛": 183, "飞": 183, "食": 184, "首": 185, "香": 186, "馬": 187,
    "骨": 188, "高": 189, "髟": 190, "鬥": 191, "鬯": 192, "鬲": 193,
    "鬼": 194, "魚": 195, "鳥": 196, "鹵": 197, "鹿": 198, "麥": 199,
    "麻": 200, "黃": 201, "黍": 202, "黑": 203, "黹": 204, "黽": 205,
    "鼎": 206, "鼓": 207, "鼠": 208, "鼻": 209, "齊": 210, "齒": 211,
    "龍": 212, "龜": 213, "龠": 214,
}


def find_desktop_xlsx() -> Path:
    one = Path("C:/Users/zairk/OneDrive")
    desk = None
    for name in os.listdir(one):
        if len(name) >= 2 and name[0] == "\u684c" and name[1] == "\u9762":
            desk = one / name
            break
    if desk is None:
        raise FileNotFoundError("Desktop directory not found")
    for f in os.listdir(desk):
        if f.startswith("CNBE") and f.endswith(".xlsx"):
            return desk / f
    raise FileNotFoundError("CNBE review xlsx not found on Desktop")


def load_radix_map(path: Path) -> dict[str, int]:
    data = json.loads(path.read_text(encoding="utf-8"))
    result: dict[str, int] = {}
    for rec in data.get("records", []):
        if rec.get("code") is not None:
            result[rec["radical"]] = rec["code"]
            if rec.get("canonical_radical"):
                result.setdefault(rec["canonical_radical"], rec["code"])
    return result


def load_unihan_krs(path: Path) -> dict[str, int]:
    result: dict[str, int] = {}
    for line in path.read_text(encoding="utf-8", errors="replace").splitlines():
        if line.startswith("#"):
            continue
        parts = line.split("\t")
        if len(parts) < 3 or parts[1] != "kRSUnicode":
            continue
        m = re.match(r"\d+", parts[2])
        if m:
            result[parts[0]] = int(m.group(0))
    return result


def main() -> None:
    xlsx_path = find_desktop_xlsx()
    radix_map = load_radix_map(REPO / "evidence" / "8105" / "cnbe8105_radical_code_map.json")
    krs = load_unihan_krs(
        EXP.parent / "2026-08-05_scheme_comparison" / "build" / "Unihan_IRGSources.txt"
    )
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[3]]

    reviewed: dict[str, dict] = {}
    unresolved: dict[str, list[str]] = {}
    for row in ws.iter_rows(min_row=2):
        c = row[0].value
        if c is None:
            continue
        radix_name = row[16].value
        strokes = row[17].value
        struct_name = row[18].value
        if radix_name is None or strokes is None or struct_name is None:
            unresolved.setdefault("missing_fields", []).append(c)
            continue
        struct_name = "独体字" if struct_name == "独体" else struct_name
        radix_code = radix_map.get(radix_name) or RADIX_SUPPLEMENT.get(radix_name)
        source = "radix_map" if radix_name in radix_map else "supplement"
        radix_name_mismatch = False
        if radix_code is None:
            ucp = f"U+{ord(c):04X}"
            if ucp in krs:
                radix_code = krs[ucp]
                source = "unihan_fallback"
                radix_name_mismatch = True
            else:
                unresolved.setdefault(radix_name, []).append(c)
        struct_code = STRUCT_CODE.get(struct_name)
        if struct_code is None:
            unresolved.setdefault(f"struct:{struct_name}", []).append(c)
        reviewed[c] = {
            "reviewed_radix_name": radix_name,
            "reviewed_strokes": int(strokes),
            "reviewed_structure": struct_name,
            "radix_code": radix_code,
            "radix_source": source,
            "radix_name_mismatch": radix_name_mismatch,
            "struct_code": struct_code,
        }

    packet = json.loads(PACKET.read_text(encoding="utf-8"))
    updated = 0
    skipped = 0
    for entry in packet["entries"]:
        if entry["action"] != "UPGRADE_CANDIDATE":
            continue
        c = entry["char"]
        if c not in reviewed:
            skipped += 1
            continue
        r = reviewed[c]
        entry["review_status"] = "APPROVED"
        entry["review"] = {
            "decision": "批准",
            "reviewer": "zairkliu",
            "reviewed_at": "2026-08-06",
            "reviewed_radix_name": r["reviewed_radix_name"],
            "reviewed_strokes": r["reviewed_strokes"],
            "reviewed_structure": r["reviewed_structure"],
            "radix_code": r["radix_code"],
            "radix_source": r["radix_source"],
            "radix_name_mismatch": r["radix_name_mismatch"],
            "struct_code": r["struct_code"],
        }
        entry["proposed"] = {
            "radix": r["radix_code"],
            "radix_name": r["reviewed_radix_name"],
            "strokes": r["reviewed_strokes"],
            "struct_name": r["reviewed_structure"],
            "struct_type": r["struct_code"],
            "index": (ord(c) - 0x4E00) % 2048,
            "ext": 0,
            "track": "provisional",
        }
        updated += 1
    packet["summary"]["batch_reviewed_approved"] = updated
    packet["summary"]["batch_skipped"] = skipped
    packet["summary"]["batch_unresolved"] = {k: len(v) for k, v in unresolved.items()}
    PACKET.write_text(json.dumps(packet, ensure_ascii=False, indent=2), encoding="utf-8")

    summary = {
        "xlsx": str(xlsx_path),
        "reviewed_rows": len(reviewed),
        "packet_updated": updated,
        "skipped": skipped,
        "unresolved": {k: len(v) for k, v in unresolved.items()},
        "unresolved_examples": {k: v[:5] for k, v in list(unresolved.items())[:10]},
    }
    (EXP / "batch_review_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
