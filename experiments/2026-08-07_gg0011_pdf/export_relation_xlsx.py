#!/usr/bin/env python3
"""Export GG0011 main/attached relation to XLSX and CSV."""

from __future__ import annotations

import csv
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

EXP = Path(__file__).resolve().parent
HEADERS = ["id", "main", "attached_forms", "ocr_main", "ocr_attached_forms", "status"]
FILL = PatternFill("solid", fgColor="2F5597")
FONT = Font(color="FFFFFF", bold=True)


def main() -> int:
    data = json.loads((EXP / "results" / "gg0011_main_attached_relation.json").read_text(encoding="utf-8"))
    rows = data["radicals"]
    csv_path = EXP / "results" / "GG0011_201_100_relation.csv"
    with open(csv_path, "w", encoding="utf-8", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(HEADERS)
        for r in rows:
            writer.writerow(
                [
                    r["id"],
                    r["main"],
                    "".join(r["attached_forms"]),
                    r["ocr_main"] or "",
                    "".join(r["ocr_attached_forms"]),
                    r["status"],
                ]
            )

    wb = Workbook()
    ws = wb.active
    ws.title = "GG0011_201_100"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.fill = FILL
        cell.font = FONT
        cell.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append(
            [
                r["id"],
                r["main"],
                "".join(r["attached_forms"]),
                r["ocr_main"] or "",
                "".join(r["ocr_attached_forms"]),
                r["status"],
            ]
        )
    for i, w in enumerate([6, 8, 16, 10, 18, 30], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    xlsx = EXP / "results" / "GG0011_201_100_relation.xlsx"
    wb.save(xlsx)
    print("saved", csv_path)
    print("saved", xlsx)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
