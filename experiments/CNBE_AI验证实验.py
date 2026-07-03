import os, json, warnings
warnings.filterwarnings("ignore")
os.environ["MPLCONFIGDIR"] = r"C:\Users\zairk\Documents\Codex\2026-06-22\ni\.matplotlib_cache"

import numpy as np
import torch
import torch.nn as nn
from torch.utils.data import TensorDataset, DataLoader
from sklearn.model_selection import train_test_split
from openpyxl import load_workbook
from collections import defaultdict

print("=" * 65)
print("CNBE AI验证 v3.0 — 编码结构深度分析")
print("=" * 65)

# ===== 加载数据 =====
print("\n[1/7] 加载CNBE编码数据...")
wb = load_workbook(r"C:\Users\zairk\Documents\Codex\2026-06-22\ni\outputs\CNBE编码目录.xlsx")
ws = wb["CNBE完整编码表"]

chars, Xc_hex, y_rad, strokes, structs = [], [], [], [], []
for row in range(2, ws.max_row + 1):
    chars.append(ws.cell(row, 2).value)
    Xc_hex.append(ws.cell(row, 4).value)   # CNBE hex
    y_rad.append(ws.cell(row, 7).value)     # 部首标签
    strokes.append(ws.cell(row, 8).value)   # 笔画数
    structs.append(ws.cell(row, 9).value)   # 结构类型

# 解析CNBE编码
Xc = np.array([int(c, 16) if isinstance(c, str) else c for c in Xc_hex], dtype=np.int64)
y_rad = np.array(y_rad, dtype=np.int64)
strokes = np.array(strokes, dtype=np.int64)
structs = np.array(structs, dtype=np.int64)

# 处理稀有部首 -> other (214)
rad_cnt = defaultdict(int)
for r in y_rad: rad_cnt[int(r)] += 1
rare = {k for k, v in rad_cnt.items() if v < 2}
y_rad_adj = y_rad.copy()
for r in rare:
    y_rad_adj[y_rad == r] = 214
n_classes_rad = 215

print(f"  样本数: {len(Xc)}, 部首类: {n_classes_rad}")

# ===== 数据集划分 =====
X_train, X_test, yr_train, yr_test, stk_train, stk_test, strc_train, strc_test = train_test_split(
    Xc, y_rad_adj, strokes, structs, test_size=0.2, random_state=42, stratify=y_rad_adj
)
print(f"\n[2/7] 数据划分: 训练 {len(X_train)}, 测试 {len(X_test)}")

# ===== 模型定义 =====
print("\n[3/7] 定义3个模型...")

class CNHEFull(nn.Module):
    """完整CNHE: 使用所有4个位域"""
    def __init__(self, n_classes):
        super().__init__()
        self.rad_emb = nn.Embedding(256, 32)
        self.stk_emb = nn.Embedding(32, 16)
        self.strc_emb = nn.Embedding(16, 16)
        self.idx_emb = nn.Embedding(2048, 32)
        self.proj = nn.Linear(96, 128)
        self.drop = nn.Dropout(0.15)
        self.norm = nn.LayerNorm(128)
        self.classifier = nn.Sequential(nn.Linear(128, 128), nn.ReLU(), nn.Dropout(0.15), nn.Linear(128, n_classes))
    def forward(self, x):
        rad = self.rad_emb((x >> 24) & 0xFF)
        stk = self.stk_emb((x >> 19) & 0x1F)
        strc = self.strc_emb((x >> 15) & 0x0F)
        idx = self.idx_emb((x >> 4) & 0x7FF)
        emb = self.proj(torch.cat([rad, stk, strc, idx], dim=-1))
        return self.classifier(self.norm(self.drop(emb)))

class CNHEMasked(nn.Module):
    """遮蔽部首: 仅使用笔画+结构+序号"""
    def __init__(self, n_classes):
        super().__init__()
        self.stk_emb = nn.Embedding(32, 48)
        self.strc_emb = nn.Embedding(16, 48)
        self.idx_emb = nn.Embedding(2048, 48)
        self.proj = nn.Linear(144, 128)
        self.drop = nn.Dropout(0.15)
        self.norm = nn.LayerNorm(128)
        self.classifier = nn.Sequential(nn.Linear(128, 128), nn.ReLU(), nn.Dropout(0.15), nn.Linear(128, n_classes))
    def forward(self, x):
        stk = self.stk_emb((x >> 19) & 0x1F)
        strc = self.strc_emb((x >> 15) & 0x0F)
        idx = self.idx_emb((x >> 4) & 0x7FF)
        emb = self.proj(torch.cat([stk, strc, idx], dim=-1))
        return self.classifier(self.norm(self.drop(emb)))

class Baseline(nn.Module):
    """Baseline: 字符ID -> Embedding"""
    def __init__(self, vocab, n_classes):
        super().__init__()
        self.emb = nn.Embedding(vocab, 128)
        self.drop = nn.Dropout(0.15)
        self.norm = nn.LayerNorm(128)
        self.classifier = nn.Sequential(nn.Linear(128, 128), nn.ReLU(), nn.Dropout(0.15), nn.Linear(128, n_classes))
    def forward(self, x):
        return self.classifier(self.norm(self.drop(self.emb(x))))

def count_params(m):
    return sum(p.numel() for p in m.parameters() if p.requires_grad)

models = {
    "Baseline": Baseline(len(Xc), n_classes_rad),
    "CNHE完整": CNHEFull(n_classes_rad),
    "CNHE屏蔽部首": CNHEMasked(n_classes_rad),
}
for name, m in models.items():
    print(f"  {name}: {count_params(m):>8,} 参数")

# ===== 训练 =====
EPOCHS, BATCH, LR = 150, 64, 5e-4
print(f"\n[4/7] 训练({EPOCHS} epochs, batch={BATCH}, lr={LR})...")

def train_model(model, X_tr, y_tr, X_te, y_te, name):
    device = "cuda" if torch.cuda.is_available() else "cpu"
    model.to(device)
    tr = DataLoader(TensorDataset(torch.tensor(X_tr, dtype=torch.long), torch.tensor(y_tr, dtype=torch.long)),
                    BATCH, shuffle=True)
    te = DataLoader(TensorDataset(torch.tensor(X_te, dtype=torch.long), torch.tensor(y_te, dtype=torch.long)),
                    BATCH)
    opt = torch.optim.AdamW(model.parameters(), lr=LR, weight_decay=1e-4)
    sched = torch.optim.lr_scheduler.CosineAnnealingLR(opt, T_max=EPOCHS)
    crit = nn.CrossEntropyLoss()
    
    hist = {"acc": [], "loss": []}
    best_acc, best_ep = 0.0, 0
    for ep in range(EPOCHS):
        model.train()
        tl = 0.0
        for Xb, yb in tr:
            Xb, yb = Xb.to(device), yb.to(device)
            opt.zero_grad()
            loss = crit(model(Xb), yb)
            loss.backward()
            nn.utils.clip_grad_norm_(model.parameters(), 1.0)
            opt.step()
            tl += loss.item()
        sched.step()
        
        model.eval()
        cor, tot = 0, 0
        with torch.no_grad():
            for Xb, yb in te:
                Xb, yb = Xb.to(device), yb.to(device)
                cor += (model(Xb).argmax(1) == yb).sum().item()
                tot += yb.size(0)
        acc = cor / tot
        hist["acc"].append(acc)
        hist["loss"].append(tl / len(tr))
        if acc > best_acc:
            best_acc, best_ep = acc, ep + 1
        if (ep + 1) % 10 == 0 or ep == 0:
            print(f"  {name} Ep {ep+1:3d}: Acc={acc:.4f}")
    print(f"  {name} Best: Ep {best_ep}, Acc={best_acc:.4f}")
    return hist, best_acc, best_ep

histories, accuracies = {}, {}
token_ids = np.arange(len(X_train))  # Baseline用

for name, model in models.items():
    if name == "Baseline":
        h, a, e = train_model(model, token_ids, yr_train, np.arange(len(X_test)), yr_test, name)
    else:
        h, a, e = train_model(model, X_train, yr_train, X_test, yr_test, name)
    histories[name] = h
    accuracies[name] = (a, e, count_params(model))

# ===== 5. 语义相似度分析 =====
print("\n[5/7] 编码相似度 vs 语义相似度分析...")

# 计算同部首 vs 跨部首的CNBE编码距离
all_chars_ord = np.argsort(X_test)
sample_size = min(500, len(X_test))

# 对每个部首类采样一对字符
within_distances = []
cross_distances = []

for i in range(200):
    # 随机选择两个字符
    idx1, idx2 = np.random.choice(len(X_test), 2, replace=False)
    c1, c2 = X_test[idx1], X_test[idx2]
    r1, r2 = yr_test[idx1], yr_test[idx2]
    ext = X_test  # ensure these are used
    # CNBE编码距离 (归一化到0-1)
    max_diff = 2 ** 32
    dist = abs(int(c1) - int(c2)) / max_diff
    
    if r1 == r2 and r1 != 214:  # 同部首且不是other
        within_distances.append(dist)
    elif r1 != r2 and r1 != 214 and r2 != 214:
        cross_distances.append(dist)

print(f"  同部首距离样本: {len(within_distances)}")
print(f"  跨部首距离样本: {len(cross_distances)}")
if within_distances and cross_distances:
    w_mean = np.mean(within_distances)
    c_mean = np.mean(cross_distances)
    w_std = np.std(within_distances)
    c_std = np.std(cross_distances)
    print(f"  同部首平均CNBE距离: {w_mean:.6f} +/- {w_std:.6f}")
    print(f"  跨部首平均CNBE距离: {c_mean:.6f} +/- {c_std:.6f}")
    print(f"  分离度(跨/同): {c_mean/w_mean:.2f}x")
    print(f"  => {'有效' if c_mean > w_mean * 1.5 else '需改进'}")

# ===== 6. 可视化 =====
print("\n[6/7] 生成对比图表...")
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

fig, axes = plt.subplots(2, 2, figsize=(14, 10))
fig.suptitle("CNBE AI验证 v3.0 — 编码结构深度分析", fontsize=14)

colors = {"Baseline": "#2F5496", "CNHE完整": "#C0504D", "CNHE屏蔽部首": "#4BACC6"}
ep_range = range(1, EPOCHS + 1)

# (1) 准确率对比
ax = axes[0, 0]
for name in ["Baseline", "CNHE完整", "CNHE屏蔽部首"]:
    ax.plot(ep_range, histories[name]["acc"], color=colors[name], label=name, linewidth=2)
ax.set_xlabel("Epoch"); ax.set_ylabel("Test Accuracy")
ax.set_title("部首分类准确率对比"); ax.legend(); ax.grid(True, alpha=0.3)

# (2) 参数量对比
ax = axes[0, 1]
names = list(accuracies.keys())
params = [accuracies[n][2] for n in names]
ax.bar(names, params, color=[colors[n] for n in names], width=0.5, edgecolor="white")
for n, p in zip(names, params):
    ax.text(names.index(n), p + max(params)*0.01, f"{p:,}\n({p/params[0]:.1%})",
            ha="center", va="bottom", fontsize=8)
ax.set_ylabel("参数量"); ax.set_title("参数量对比")
ax.set_xticklabels(names, fontsize=9)

# (3) 距离分布直方图
ax = axes[1, 0]
if within_distances and cross_distances:
    ax.hist(within_distances, bins=30, alpha=0.6, color="#4BACC6", label=f"同部首(mean={w_mean:.4f})")
    ax.hist(cross_distances, bins=30, alpha=0.6, color="#C0504D", label=f"跨部首(mean={c_mean:.4f})")
    ax.axvline(w_mean, color="#4BACC6", ls="--", lw=2)
    ax.axvline(c_mean, color="#C0504D", ls="--", lw=2)
ax.set_xlabel("CNBE编码距离"); ax.set_ylabel("频次"); ax.set_title("CNBE编码距离分布")
ax.legend(); ax.grid(True, alpha=0.3)

# (4) 最终准确率柱状图
ax = axes[1, 1]
acc_vals = [accuracies[n][0] * 100 for n in names]
bars = ax.bar(names, acc_vals, color=[colors[n] for n in names], width=0.5, edgecolor="white")
for bar, val in zip(bars, acc_vals):
    ax.text(bar.get_x()+bar.get_width()/2, bar.get_height()+1,
            f"{val:.1f}%", ha="center", va="bottom", fontsize=9)
ax.set_ylabel("准确率(%)"); ax.set_title("最终准确率对比")
ax.set_xticklabels(names, fontsize=9)
ax.set_ylim(0, 110)

plt.tight_layout()
chart = r"C:\Users\zairk\Documents\Codex\2026-06-22\ni\experiment\outputs\cnbe_v3_analysis.png"
plt.savefig(chart, dpi=150, bbox_inches="tight")
plt.close()
print(f"  图表: {chart}")

# ===== 7. 报告 =====
print("\n[7/7] 生成报告...")

# 收敛分析
def conv_epoch(h, target_pct=0.95):
    best = max(h["acc"])
    target = best * target_pct
    eps = [i+1 for i, a in enumerate(h["acc"]) if a >= target]
    return eps[0] if eps else EPOCHS

report = {}
for n in names:
    a, e, p = accuracies[n]
    report[n] = {
        "accuracy": float(a),
        "best_epoch": e,
        "params": p,
        "convergence_epoch": conv_epoch(histories[n]),
    }

report["distance_analysis"] = {
    "within_same_radical": {
        "mean": float(np.mean(within_distances)) if within_distances else 0,
        "std": float(np.std(within_distances)) if within_distances else 0,
        "samples": len(within_distances),
    },
    "cross_radical": {
        "mean": float(np.mean(cross_distances)) if cross_distances else 0,
        "std": float(np.std(cross_distances)) if cross_distances else 0,
        "samples": len(cross_distances),
    }
}

rpath = r"C:\Users\zairk\Documents\Codex\2026-06-22\ni\experiment\outputs\v3_report.json"
with open(rpath, "w", encoding="utf-8") as f:
    json.dump(report, f, ensure_ascii=False, indent=2)

print("\n" + "=" * 65)
print("实验结果汇总")
print("=" * 65)
for n in names:
    a, e, p = accuracies[n]
    conv = conv_epoch(histories[n])
    print(f"\n  {n}:")
    print(f"    准确率: {a:.2%} (Epoch {e})")
    print(f"    收敛:   {conv} Epoch")
    print(f"    参数:   {p:,}")
print(f"\n  编码距离分析:")
print(f"    同部首字 CNBE距离: {np.mean(within_distances):.6f}" if within_distances else "")
print(f"    跨部首字 CNBE距离: {np.mean(cross_distances):.6f}" if cross_distances else "")
print(f"    分离度:            {np.mean(cross_distances)/np.mean(within_distances):.2f}x" if within_distances and cross_distances else "")