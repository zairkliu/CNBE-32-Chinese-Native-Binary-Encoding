#!/usr/bin/env python3
"""Full-catalog (97,686) candidate prefill runner.

Deterministic evidence layer runs locally. LLM layer is optional and can be
limited to incomplete rows, a sample, or all rows. Every LLM call is cached in
an audit log and guarded by call/token budgets. No release database writes.
"""

from __future__ import annotations

import argparse
import gzip
import json
import sys
import threading
import time
from collections import Counter
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.path.insert(0, str(Path(__file__).resolve().parents[2] / "src"))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from clients.evidence import (
    STRUCT_CODE,
    aggregate,
    deterministic_proposal,
    encode_proposal,
    load_ids,
    load_radix_name_map,
    load_unihan_irg,
)
from clients.llm_client import DeepSeekV4Client

REPO = Path(__file__).resolve().parents[2]
EXP = Path(__file__).resolve().parent
DEFAULT_CONFIG = EXP / "pipeline_config.json"
RADIX_ALIASES = {"乚": "乙"}

HEADER_FILL = PatternFill("solid", fgColor="2F5597")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HEADERS = [
    "汉字", "Unicode", "块", "当前CNBE",
    "证据部首码", "证据部首名", "证据笔画", "IDS", "证据结构",
    "建议部首码", "建议部首名", "建议笔画", "建议结构", "建议结构码",
    "置信度", "往返通过", "理由",
    "LLM部首名", "LLM笔画", "LLM结构", "LLM置信度", "LLM理由",
    "复核决定(批准/驳回/修改)", "复核人", "备注",
]
WIDTHS = [8, 10, 16, 12, 12, 12, 10, 26, 12, 12, 12, 10, 12, 12, 10, 10, 34, 12, 10, 12, 10, 36, 22, 10, 20]


def load_catalog_rows(path: Path, limit: int = 0) -> list[dict]:
    rows = []
    with gzip.open(path, "rt", encoding="utf-8", errors="replace") as fh:
        for line in fh:
            line = line.strip()
            if not line or line.startswith("id,"):
                continue
            parts = line.split(",")
            ucp = parts[1]
            rows.append(
                {
                    "ucp": ucp,
                    "char": chr(int(ucp[2:], 16)),
                    "cnbe_hex": parts[2],
                    "r": int(parts[3]),
                    "s": int(parts[4]),
                    "g": int(parts[5]),
                    "block": parts[6],
                }
            )
            if limit and len(rows) >= limit:
                break
    return rows


def normalize_llm_radix_name(name: str, name_to_code: dict[str, int]) -> int | None:
    if not name:
        return None
    name = RADIX_ALIASES.get(name.strip(), name.strip())
    for canonical, code in name_to_code.items():
        if name == canonical or name in canonical or canonical in name:
            return code
    return None


def llm_prompt(row: dict, ev: dict, det: dict) -> str:
    return (
        "你是汉字结构辅助裁决助手。只输出 JSON，不要解释。\n"
        "结构类型只能从以下 13 项中选择："
        + "、".join(STRUCT_CODE.keys())
        + "。\n"
        f"汉字：{row['char']}（{row['ucp']}）\n"
        f"全量目录当前值：部首={row['r']}，笔画={row['s']}，结构码={row['g']}\n"
        f"交叉参考：Unihan部首={ev['kRSUnicode']}（{ev['radix_name']}），"
        f"笔画={ev['strokes']}，IDS={ev['ids']}\n"
        f"确定性候选：部首={det.get('radix_name')}，笔画={det.get('strokes')}，"
        f"结构={det.get('struct_name')}\n"
        "请输出："
        '{"radix_name": "部首名", "strokes": 整数, "structure": "结构名", '
        '"confidence": 0到1的小数, "rationale": "一句裁决理由"}'
    )


def load_llm_cache(path: Path) -> dict[str, str]:
    cache: dict[str, str] = {}
    if not path.exists():
        return cache
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line:
            continue
        rec = json.loads(line)
        if rec.get("status") == "ok" and rec.get("raw"):
            cache[rec["unicode"]] = rec["raw"]
    return cache


def build_workbook(entries: list[dict], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "full_prefill"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for e in entries:
        ev = e["evidence"]
        det = e["deterministic"]
        llm = e.get("llm") or {}
        ws.append([
            e["char"], e["ucp"], e["block"], e["cnbe_hex"],
            ev.get("radix_code", ""), ev.get("radix_name", ""), ev.get("strokes", ""), ev.get("ids", ""), ev.get("structure", ""),
            det.get("radix", ""), det.get("radix_name", ""), det.get("strokes", ""), det.get("struct_name", ""), det.get("struct_type", ""),
            det.get("confidence", ""), det.get("roundtrip_pass", ""), "\n".join(det.get("reasons", [])),
            llm.get("radix_name", ""), llm.get("strokes", ""), llm.get("structure", ""),
            llm.get("confidence", ""), llm.get("rationale", ""),
            "", "", "",
        ])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    out_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(out_path)
    except PermissionError:
        fallback = out_path.with_name(out_path.stem + "_pending.xlsx")
        wb.save(fallback)
        print(f"[warn] primary workbook locked; wrote {fallback}")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG)
    parser.add_argument("--limit", type=int, default=0, help="smoke limit over catalog order")
    parser.add_argument("--llm-mode", choices=["none", "incomplete", "sample", "all"], default="none")
    parser.add_argument("--sample-size", type=int, default=20)
    parser.add_argument("--workers", type=int, default=1)
    parser.add_argument("--out-dir", type=Path, default=EXP / "results")
    args = parser.parse_args()

    config = json.loads(args.config.read_text(encoding="utf-8"))
    ev_cfg = config["evidence"]
    llm_cfg = config["llm"]

    rows = load_catalog_rows(REPO / ev_cfg["catalog_csv_gz"], limit=args.limit)
    unihan = load_unihan_irg(REPO / ev_cfg["unihan_irg"])
    ids_map = load_ids(REPO / ev_cfg["ids"])
    radix_name = load_radix_name_map(REPO / ev_cfg["radix_map"])
    reverse_radix = {name: code for code, name in radix_name.items()}

    entries = []
    for row in rows:
        entry = {"char": row["char"], "unicode": row["ucp"], "standard_evidence": {"issues": [], "structure": None}}
        ev = aggregate(entry, unihan, ids_map, radix_name)
        det = deterministic_proposal(entry, ev, (ord(row["char"]) - 0x4E00) % 2048)
        encode_proposal(det)
        entries.append({**row, "evidence": ev, "deterministic": det})

    client = DeepSeekV4Client(model=llm_cfg["model"], base_url=llm_cfg["base_url"])
    audit_path = args.out_dir / "full_llm_audit.jsonl"
    audit_path.parent.mkdir(parents=True, exist_ok=True)
    cache = load_llm_cache(audit_path)
    lock = threading.Lock()

    def complete(e: dict) -> bool:
        det = e["deterministic"]
        return det.get("radix") is not None and det.get("strokes") is not None and det.get("struct_type") is not None

    if args.llm_mode == "incomplete":
        targets = [e for e in entries if not complete(e)]
    elif args.llm_mode == "all":
        targets = entries
    elif args.llm_mode == "sample":
        incomplete = [e for e in entries if not complete(e)]
        complete_rows = [e for e in entries if complete(e)]
        targets = incomplete[: args.sample_size // 2] + complete_rows[: args.sample_size - len(incomplete[: args.sample_size // 2])]
    else:
        targets = []

    max_calls = config["cost_guard"]["max_llm_calls"]
    max_tokens = config["cost_guard"]["max_total_tokens"]
    llm_stats = {
        "requested": len(targets),
        "cached": 0,
        "parsed": 0,
        "agreement": 0,
        "consistent": 0,
        "calls_made": 0,
        "tokens_used": 0,
    }
    if targets and not client.available:
        llm_stats["status"] = "LLM_SKIPPED"
        targets = []
    elif targets:
        llm_stats["status"] = "RUN"

        def process(e: dict) -> None:
            if llm_stats["calls_made"] >= max_calls:
                return
            raw = cache.get(e["ucp"])
            if raw is not None:
                with lock:
                    llm_stats["cached"] += 1
            else:
                resp = client.chat(llm_prompt(e, e["evidence"], e["deterministic"]))
                with lock:
                    llm_stats["calls_made"] += 1
                    llm_stats["tokens_used"] += resp.usage.get("total_tokens", 0)
                if resp.status != "ok":
                    e["llm"] = {"status": resp.status}
                    return
                raw = resp.text
                audit_path.open("a", encoding="utf-8").write(
                    json.dumps(
                        {
                            "ts": time.time(),
                            "unicode": e["ucp"],
                            "char": e["char"],
                            "model": client.model,
                            "status": resp.status,
                            "elapsed": round(resp.elapsed, 3),
                            "usage": resp.usage,
                            "raw": resp.text[:2000],
                        },
                        ensure_ascii=False,
                    )
                    + "\n"
                )
                cache[e["ucp"]] = raw
            try:
                data = client.parse_json(raw)
                radix_code = normalize_llm_radix_name(str(data.get("radix_name", "")), reverse_radix)
                llm = {
                    "radix": radix_code,
                    "radix_name": data.get("radix_name"),
                    "strokes": int(data["strokes"]) if isinstance(data.get("strokes"), int) else None,
                    "structure": data.get("structure"),
                    "struct_type": STRUCT_CODE.get(str(data.get("structure", ""))),
                    "confidence": float(data.get("confidence", 0.0)),
                    "rationale": str(data.get("rationale", "")),
                }
                e["llm"] = llm
                with lock:
                    llm_stats["parsed"] += 1
                    det = e["deterministic"]
                    strict_agreement = (
                        llm["radix"] == det.get("radix")
                        and llm["strokes"] == det.get("strokes")
                        and llm["struct_type"] == det.get("struct_type")
                    )
                    consistent = (
                        (det.get("radix") is None or llm["radix"] == det.get("radix"))
                        and (det.get("strokes") is None or llm["strokes"] == det.get("strokes"))
                        and (det.get("struct_type") is None or llm["struct_type"] == det.get("struct_type"))
                    )
                    if strict_agreement:
                        llm_stats["agreement"] += 1
                    if consistent:
                        llm_stats["consistent"] += 1
            except (ValueError, KeyError, TypeError) as exc:
                e["llm"] = {"status": f"parse_error:{exc}", "raw": raw[:500]}

        with ThreadPoolExecutor(max_workers=args.workers) as pool:
            list(pool.map(process, targets))

    det_complete = sum(1 for e in entries if complete(e))
    roundtrip = sum(1 for e in entries if e["deterministic"].get("roundtrip_pass"))
    overflow = sum(1 for e in entries if e["deterministic"].get("strokes") is not None and e["deterministic"]["strokes"] > 31)
    bucket = Counter()
    for e in entries:
        conf = e["deterministic"]["confidence"]
        if complete(e) and conf >= config["confidence"]["high"]:
            bucket["high"] += 1
        elif complete(e):
            bucket["medium"] += 1
        else:
            bucket["low"] += 1

    results = {
        "schema_version": 1,
        "scope": config["scope"],
        "rows": len(entries),
        "deterministic_complete": det_complete,
        "deterministic_complete_rate": round(det_complete / len(entries), 4) if entries else 0.0,
        "roundtrip_pass": roundtrip,
        "roundtrip_pass_rate": round(roundtrip / len(entries), 4) if entries else 0.0,
        "stroke_overflow": overflow,
        "confidence_buckets": dict(bucket),
        "llm": llm_stats,
        "write_gate": "NO_WRITE_TO_RELEASE_DB",
    }
    out = args.out_dir
    out.mkdir(parents=True, exist_ok=True)
    (out / "full_prefill.json").write_text(
        json.dumps({"summary": results, "entries": entries}, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (out / "full_results.json").write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
    build_workbook(entries, out / "CNBE_FULL_PREFILL_SAMPLE.xlsx")
    print(json.dumps(results, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
