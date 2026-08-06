#!/usr/bin/env python3
"""Prefill the 491 remaining legacy rows with API/evidence candidates.

Layer 1 is deterministic: Unihan + CHISE IDS + CNBE radix map. Layer 2 is an
optional DeepSeek V4 call that returns a candidate and confidence. Outputs are
review packets only; this script never writes the release database.
"""

from __future__ import annotations

import argparse
import json
import sys
import time
from collections import Counter
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[2] / "src"))
sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from cnbe32 import decode_cnbe, encode_cnbe

from clients.evidence import (
    STRUCT_CODE,
    aggregate,
    deterministic_proposal,
    encode_proposal,
    load_ids,
    load_radix_name_map,
    load_unihan_irg,
)
from clients.llm_client import DeepSeekV4Client

REPO = Path(__file__).resolve().parents[2]
EXP = Path(__file__).resolve().parent
RADIX_ALIASES = {"乚": "乙"}

HEADER_FILL = PatternFill("solid", fgColor="2F5597")
HEADER_FONT = Font(color="FFFFFF", bold=True)

HEADERS = [
    "汉字", "Unicode", "8105序号",
    "当前部首码", "当前部首名", "当前笔画", "当前结构",
    "证据部首码", "证据部首名", "证据笔画", "IDS", "证据结构",
    "建议部首码", "建议部首名", "建议笔画", "建议结构", "建议结构码",
    "置信度", "理由",
    "LLM部首名", "LLM笔画", "LLM结构", "LLM置信度", "LLM理由",
    "复核决定(批准/驳回/修改)", "复核人", "备注",
]
WIDTHS = [8, 10, 10, 10, 10, 10, 12, 12, 12, 10, 24, 12, 12, 12, 10, 12, 12, 10, 34, 12, 10, 12, 10, 36, 22, 10, 20]


def llm_prompt(entry: dict, ev: dict, proposal: dict) -> str:
    std = entry.get("standard_evidence", {})
    return (
        "你是汉字结构辅助裁决助手。只输出 JSON，不要解释。\n"
        "结构类型只能从以下 13 项中选择："
        + "、".join(STRUCT_CODE.keys())
        + "。\n"
        f"汉字：{entry['char']}（{entry['unicode']}）\n"
        f"当前库字段：部首={entry['current'].get('radix_name')}，"
        f"笔画={entry['current'].get('strokes')}，结构={entry['current'].get('struct_name')}\n"
        f"标准证据：部首={std.get('radical')}，笔画={std.get('stroke_count')}，"
        f"结构={std.get('structure')}，分解={std.get('decomposition')}\n"
        f"交叉参考：Unihan部首={ev['kRSUnicode']}（{ev['radix_name']}），"
        f"笔画={ev['strokes']}，IDS={ev['ids']}\n"
        f"确定性候选：部首={proposal.get('radix_name')}，笔画={proposal.get('strokes')}，"
        f"结构={proposal.get('struct_name')}\n"
        "请输出："
        '{"radix_name": "部首名", "strokes": 整数, "structure": "结构名", '
        '"confidence": 0到1的小数, "rationale": "一句裁决理由"}'
    )


def normalize_llm_radix_name(name: str, name_to_code: dict[str, int]) -> int | None:
    if not name:
        return None
    name = name.strip()
    name = RADIX_ALIASES.get(name, name)
    for canonical, code in name_to_code.items():
        if name == canonical or name in canonical or canonical in name:
            return code
    return None


def load_llm_cache(path: Path) -> dict[str, str]:
    cache: dict[str, str] = {}
    if not path.exists():
        return cache
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line:
            continue
        rec = json.loads(line)
        if rec.get("status") == "ok" and rec.get("raw"):
            cache[rec["char"]] = rec["raw"]
    return cache


def build_workbook(entries: list[dict], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "legacy491_prefill"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for e in entries:
        cur = e["current"]
        ev = e["evidence"]
        det = e["deterministic"]
        llm = e.get("llm") or {}
        ws.append([
            e["char"], e["unicode"], e.get("standard_rank") or "",
            cur.get("radix", ""), cur.get("radix_name", ""), cur.get("strokes", ""), cur.get("struct_name", ""),
            ev.get("radix_code", ""), ev.get("radix_name", ""), ev.get("strokes", ""), ev.get("ids", ""), ev.get("structure", ""),
            det.get("radix", ""), det.get("radix_name", ""), det.get("strokes", ""), det.get("struct_name", ""), det.get("struct_type", ""),
            det.get("confidence", ""), "\n".join(det.get("reasons", [])),
            llm.get("radix_name", ""), llm.get("strokes", ""), llm.get("structure", ""),
            llm.get("confidence", ""), llm.get("rationale", ""),
            "", "", "",
        ])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    out_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(out_path)
    except PermissionError:
        fallback = out_path.with_name(out_path.stem + "_pending.xlsx")
        wb.save(fallback)
        print(f"[warn] primary workbook locked; wrote {fallback}")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--packet", type=Path, default=REPO / "evidence/8105/8105_REMAINING_503_COMPLETION_PACKET.json")
    parser.add_argument("--unihan", type=Path, default=REPO / "experiments/2026-08-05_scheme_comparison/build/Unihan_IRGSources.txt")
    parser.add_argument("--ids", type=Path, default=REPO / "experiments/2026-08-05_scheme_comparison/build/ids.txt")
    parser.add_argument("--radix-map", type=Path, default=REPO / "evidence/8105/cnbe8105_radical_code_map.json")
    parser.add_argument("--out-dir", type=Path, default=EXP / "results")
    parser.add_argument("--llm-limit", type=int, default=0)
    parser.add_argument("--llm-model", type=str, default="deepseek-v4-flash")
    args = parser.parse_args()

    packet = json.loads(args.packet.read_text(encoding="utf-8"))
    entries = packet["entries"]
    unihan = load_unihan_irg(args.unihan)
    ids_map = load_ids(args.ids)
    radix_name = load_radix_name_map(args.radix_map)
    reverse_radix = {name: code for code, name in radix_name.items()}

    client = DeepSeekV4Client(model=args.llm_model)
    audit_path = args.out_dir / "llm_audit.jsonl"
    audit_path.parent.mkdir(parents=True, exist_ok=True)
    llm_cache = load_llm_cache(audit_path)

    enriched = []
    for entry in entries:
        ev = aggregate(entry, unihan, ids_map, radix_name)
        det = deterministic_proposal(entry, ev, (ord(entry["char"]) - 0x4E00) % 2048)
        encode_proposal(det)
        enriched.append({"char": entry["char"], "unicode": entry["unicode"], "standard_rank": entry.get("standard_rank"), "current": entry["current"], "evidence": ev, "deterministic": det, "entry": entry})

    llm_targets = enriched[: args.llm_limit] if args.llm_limit > 0 else []
    llm_stats = {"requested": len(llm_targets), "cached": 0, "parsed": 0, "agreement": 0, "consistent": 0}
    if llm_targets and not client.available:
        print("LLM_SKIPPED: no API key configured", flush=True)
        llm_stats["status"] = "LLM_SKIPPED"
    elif llm_targets:
        llm_stats["status"] = "RUN"
        for item in llm_targets:
            prompt = llm_prompt(item["entry"], item["evidence"], item["deterministic"])
            raw_text = llm_cache.get(item["char"])
            if raw_text is not None:
                llm_stats["cached"] += 1
            else:
                resp = client.chat(prompt)
                audit_path.open("a", encoding="utf-8").write(
                    json.dumps(
                        {
                            "ts": time.time(),
                            "char": item["char"],
                            "unicode": item["unicode"],
                            "model": client.model,
                            "status": resp.status,
                            "elapsed": round(resp.elapsed, 3),
                            "usage": resp.usage,
                            "raw": resp.text[:2000],
                        },
                        ensure_ascii=False,
                    )
                    + "\n"
                )
                if resp.status != "ok":
                    item["llm"] = {"status": resp.status}
                    continue
                raw_text = resp.text
            try:
                data = client.parse_json(raw_text)
                radix_code = normalize_llm_radix_name(str(data.get("radix_name", "")), reverse_radix)
                llm = {
                    "radix": radix_code,
                    "radix_name": data.get("radix_name"),
                    "strokes": int(data["strokes"]) if isinstance(data.get("strokes"), int) else None,
                    "structure": data.get("structure"),
                    "struct_type": STRUCT_CODE.get(str(data.get("structure", ""))),
                    "confidence": float(data.get("confidence", 0.0)),
                    "rationale": str(data.get("rationale", "")),
                }
                item["llm"] = llm
                llm_stats["parsed"] += 1
                det = item["deterministic"]
                strict_agreement = (
                    llm["radix"] == det.get("radix")
                    and llm["strokes"] == det.get("strokes")
                    and llm["struct_type"] == det.get("struct_type")
                )
                consistent = (
                    (det.get("radix") is None or llm["radix"] == det.get("radix"))
                    and (det.get("strokes") is None or llm["strokes"] == det.get("strokes"))
                    and (det.get("struct_type") is None or llm["struct_type"] == det.get("struct_type"))
                )
                if (
                    strict_agreement
                ):
                    llm_stats["agreement"] += 1
                if consistent:
                    llm_stats["consistent"] += 1
            except (ValueError, KeyError, TypeError) as exc:
                item["llm"] = {"status": f"parse_error:{exc}", "raw": resp.text[:500]}

    det_complete = sum(
        1 for e in enriched
        if e["deterministic"].get("radix") is not None
        and e["deterministic"].get("strokes") is not None
        and e["deterministic"].get("struct_type") is not None
    )
    roundtrip_pass = sum(1 for e in enriched if e["deterministic"].get("roundtrip_pass"))
    confidence_avg = round(
        sum(e["deterministic"]["confidence"] for e in enriched) / len(enriched), 4
    ) if enriched else 0.0
    issue_counter = Counter(
        issue
        for e in enriched
        for issue in e["entry"].get("standard_evidence", {}).get("issues", [])
    )

    results = {
        "schema_version": 1,
        "generated_at": time.strftime("%Y-%m-%dT%H:%M:%S%z"),
        "rows": len(enriched),
        "deterministic_complete": det_complete,
        "deterministic_complete_rate": round(det_complete / len(enriched), 4) if enriched else 0.0,
        "roundtrip_pass": roundtrip_pass,
        "roundtrip_pass_rate": round(roundtrip_pass / len(enriched), 4) if enriched else 0.0,
        "confidence_avg": confidence_avg,
        "issue_counts": dict(issue_counter),
        "llm": llm_stats,
        "write_gate": "NO_WRITE_TO_RELEASE_DB",
    }

    out_dir = args.out_dir
    out_dir.mkdir(parents=True, exist_ok=True)
    payload = {"summary": results, "entries": enriched}
    (out_dir / "legacy491_prefill.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (out_dir / "results.json").write_text(
        json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    xlsx = out_dir / "CNBE8105_LEGACY_PREFILL_2026-08-07.xlsx"
    build_workbook(enriched, xlsx)

    report = [
        "# 8105 Legacy 491 预填实验（2026-08-07）",
        "",
        "## 层 1：确定性证据预填（全部 491 行）",
        "",
        f"- 完整候选（部首/笔画/结构均可用）：{det_complete}/{len(enriched)} "
        f"({results['deterministic_complete_rate']})",
        f"- CNBE encode/decode 往返通过：{roundtrip_pass}/{len(enriched)} "
        f"({results['roundtrip_pass_rate']})",
        f"- 平均置信度：{confidence_avg}",
        "",
        "## 层 2：LLM API 预填",
        "",
        f"- 请求：{llm_stats.get('requested', 0)}，状态：{llm_stats.get('status', 'SKIPPED')}，"
        f"解析：{llm_stats.get('parsed', 0)}，与确定性候选一致：{llm_stats.get('agreement', 0)}",
        "",
        "## 结论与边界",
        "",
        "- 本实验只生成候选与置信度，不构成国标锚定。",
        "- 人工审核仍以独立工作簿为准；预填表用于缩短人工裁决时间。",
        "- 完整输出：`outputs/legacy491_prefill.json`、`outputs/CNBE8105_LEGACY_PREFILL_2026-08-07.xlsx`。",
        "",
    ]
    (out_dir / "REPORT.md").write_text("\n".join(report), encoding="utf-8")
    print(json.dumps(results, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
