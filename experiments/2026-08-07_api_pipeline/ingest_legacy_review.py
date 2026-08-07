#!/usr/bin/env python3
"""Ingest the human-reviewed legacy prefill workbook into a reviewed packet.

The workbook edits the suggestion columns instead of the decision column.
Unchanged rows are treated as approved; changed rows are REVIEWED_MODIFIED.
Radix codes are resolved from a small human-review override table plus the
project radix map. No release database write is performed.
"""

from __future__ import annotations

import argparse
import json
import sys
import time
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[2] / "src"))

from cnbe32 import decode_cnbe, encode_cnbe  # noqa: E402

REPO = Path(__file__).resolve().parents[2]
EXP = Path(__file__).resolve().parent

# Human-review radix name overrides (reviewed name -> project radix code).
RADIX_OVERRIDES = {
    "王": 96,
    "衤": 145,
    "罒": 122,
    "月": 74,
    "长": 168,
    "飞": 183,
    "乙": 5,
    "母": 80,
}

STRUCT_ALIASES = {"独体": "独体字", "独体字": "独体字"}
STRUCT_CODES = {
    "独体字": 0, "上下": 1, "上中下": 2, "左右": 3, "左中右": 4,
    "左上包": 5, "右上包": 6, "左三包": 7, "左下包": 8,
    "上三包": 9, "下三包": 10, "全包围": 11, "镶嵌": 12,
}


def load_workbook_rows(path: Path) -> list[tuple]:
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    return list(ws.iter_rows(min_row=2, values_only=True))


def encode_candidate(radix_code: int | None, strokes: int | None, struct_code: int | None, index: int) -> dict:
    if any(v is None for v in (radix_code, strokes, struct_code)):
        return {"cnbe": None, "roundtrip_pass": False}
    try:
        code = encode_cnbe(int(radix_code), int(strokes), int(struct_code), int(index), 0)
        decoded = decode_cnbe(code)
        return {
            "cnbe": code.code,
            "cnbe_hex": hex(code.code),
            "roundtrip_pass": (
                decoded["radix"] == int(radix_code)
                and decoded["stroke"] == int(strokes)
                and decoded["struct"] == int(struct_code)
                and decoded["index"] == int(index)
            ),
        }
    except Exception as exc:
        return {"cnbe": None, "roundtrip_pass": False, "error": str(exc)}


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=None, help="default: find under OneDrive")
    parser.add_argument("--prefill-json", type=Path, default=EXP / "results/legacy491_prefill.json")
    parser.add_argument("--out-dir", type=Path, default=EXP / "results")
    args = parser.parse_args()

    workbook = args.workbook
    if workbook is None:
        candidates = list(Path(r"C:/Users/zairk/OneDrive").rglob("CNBE8105_LEGACY_PREFILL_2026-08-07.xlsx"))
        if not candidates:
            raise SystemExit("workbook not found under OneDrive")
        workbook = candidates[0]

    rows = load_workbook_rows(workbook)
    prefill = json.loads(args.prefill_json.read_text(encoding="utf-8"))["entries"]
    by_ucp = {e["unicode"]: e for e in prefill}

    entries = []
    stats = Counter()
    radix_changes: Counter[tuple[str, str]] = Counter()
    struct_changes: Counter[tuple[str, str]] = Counter()
    unresolved = []
    for r in rows:
        ucp = r[1]
        source = by_ucp[ucp]
        det = source["deterministic"]
        reviewed = {
            "radix_name": r[13],
            "strokes": r[14],
            "struct_name": r[15],
            "struct_code": r[16],
            "confidence": r[17],
            "note": r[26] or "",
        }
        modified = (
            reviewed["radix_name"] != det.get("radix_name")
            or reviewed["strokes"] != det.get("strokes")
            or reviewed["struct_name"] != det.get("struct_name")
        )
        status = "REVIEWED_MODIFIED" if modified else "REVIEWED_APPROVED"
        stats[status] += 1

        if reviewed["radix_name"] != det.get("radix_name") and reviewed["radix_name"]:
            key = (str(det.get("radix_name")), str(reviewed["radix_name"]))
            radix_changes[key] += 1
        if reviewed["struct_name"] != det.get("struct_name") and reviewed["struct_name"]:
            key = (str(det.get("struct_name")), str(reviewed["struct_name"]))
            struct_changes[key] += 1

        struct_name = STRUCT_ALIASES.get(reviewed["struct_name"], reviewed["struct_name"])
        struct_code = STRUCT_CODES.get(struct_name, reviewed["struct_code"])
        radix_name = reviewed["radix_name"]
        radix_code = RADIX_OVERRIDES.get(radix_name) if radix_name else None
        if radix_code is None and radix_name == det.get("radix_name"):
            radix_code = det.get("radix")
        if radix_code is None:
            unresolved.append(ucp)
        proposed = {
            "radix": radix_code,
            "radix_name": radix_name,
            "strokes": reviewed["strokes"],
            "struct_name": struct_name,
            "struct_type": struct_code,
            "index": det.get("index"),
            "ext": 0,
            "track": "provisional",
        }
        proposed.update(encode_candidate(radix_code, reviewed["strokes"], struct_code, det.get("index")))
        stats["roundtrip_pass"] += int(proposed.get("roundtrip_pass") is True)
        entries.append(
            {
                "char": r[0],
                "unicode": ucp,
                "standard_rank": r[2],
                "current": source["current"],
                "deterministic": det,
                "reviewed": reviewed,
                "review_status": status,
                "proposed": proposed,
            }
        )

    packet = {
        "schema_version": 1,
        "reviewed_at": time.strftime("%Y-%m-%dT%H:%M:%S%z"),
        "workbook": str(workbook),
        "summary": {
            "total": len(entries),
            "approved": stats["REVIEWED_APPROVED"],
            "modified": stats["REVIEWED_MODIFIED"],
            "roundtrip_pass": stats["roundtrip_pass"],
            "radix_changes": {f"{k[0]}->{k[1]}": v for k, v in radix_changes.items()},
            "struct_changes": {f"{k[0]}->{k[1]}": v for k, v in struct_changes.items()},
            "unresolved_radix_code": unresolved,
        },
        "entries": entries,
    }
    out = args.out_dir
    out.mkdir(parents=True, exist_ok=True)
    (out / "legacy491_reviewed_packet.json").write_text(
        json.dumps(packet, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(json.dumps(packet["summary"], ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
