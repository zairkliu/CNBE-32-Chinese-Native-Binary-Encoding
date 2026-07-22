#!/usr/bin/env python3
"""
CNBE-32 编码表生成器 v2.0 — 排序表 + 哈希自索引
功能:
  1. 从 Excel 读取已验证的 CNBE 编码
  2. 按 Unicode 排序 → 支持二分查找 (312x加速)
  3. 生成 15bit 哈希索引表 → 支持 O(1) 查找 (4052x加速)
  4. 碰撞分析: 最大碰撞 2, 99.8% 直接命中
"""
import os, math

# Configure via CNBE_SOURCE_XLSX env var
_DEFAULT_XLSX = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data", "cnbe_catalog.xlsx")
XLSX = os.environ.get("CNBE_SOURCE_XLSX", _DEFAULT_XLSX)
if not os.path.exists(XLSX):
    raise FileNotFoundError(
        f"Source XLSX not found: {XLSX}.\n"
        "Set CNBE_SOURCE_XLSX env var to your encoding catalog file, or\n"
        f"place the file at {_DEFAULT_XLSX}"
    )
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)))

def main():
    from openpyxl import load_workbook
    wb = load_workbook(XLSX, read_only=True)
    ws = wb.active

    # 加载原始数据
    pairs = []  # (unicode, cnbe_code, char)
    for row in ws.iter_rows(min_row=2, max_row=8106, values_only=True):
        _, ch, us, hex_str, *_ = row
        if ch and isinstance(us, str) and us.startswith("U+") and hex_str:
            pairs.append((int(us[2:], 16), int(hex_str, 16), ch))
    wb.close()
    print(f"加载 {len(pairs)} 个汉字")

    # 验证唯一性
    assert len({c for _, c, _ in pairs}) == 8105
    assert len({u for u, _, _ in pairs}) == 8105
    print("编码唯一性: ✓ (8105/8105)")

    # ============================================================
    # 1. 按 Unicode 排序 → 支持二分查找
    # ============================================================
    sorted_pairs = sorted(pairs, key=lambda x: x[0])
    unicodes  = [u for u, _, _ in sorted_pairs]
    cnbe_codes = [c for _, c, _ in sorted_pairs]
    chars     = [ch for _, _, ch in sorted_pairs]

    # 写入排序后的 cnbe_table.h
    with open(os.path.join(OUT, "cnbe_table.h"), "w", encoding="utf-8") as f:
        f.write("// CNBE-32 Encoding Table (Sorted by Unicode)\n")
        f.write("// Auto-generated from CNBE V6.0 validated encoding catalog\n")
        f.write("// Sorted for O(log n) binary search (13 comparisons vs 4053)\n")
        f.write("#ifndef CNBE_TABLE_H\n#define CNBE_TABLE_H\n#include <stdint.h>\n\n")
        f.write("static const uint32_t cnbe_table[8105] = {\n")
        for i, code in enumerate(cnbe_codes):
            rad = (code>>24)&0xFF; st = (code>>19)&0x1F; sg = (code>>15)&0xF
            ix = (code>>4)&0x7FF; ex = code&0xF
            f.write(f"    0x{code:08X}, // [{i:4d}] U+{unicodes[i]:04X} '{chars[i]}'"
                    f" R={rad:3d} S={st:2d} G={sg:2d} I={ix:4d} E={ex:2d}\n")
        f.write("};\n\n")
        f.write("/* Binary search: Unicode → CNBE */\n")
        f.write("static inline uint32_t cnbe_enc_bsearch(uint32_t ucp) {\n")
        f.write("    int lo = 0, hi = 8104;\n")
        f.write("    while (lo <= hi) {\n")
        f.write("        int mid = (lo + hi) >> 1;\n")
        f.write("        if (unicode_table[mid] == ucp) return cnbe_table[mid];\n")
        f.write("        if (unicode_table[mid] < ucp) lo = mid + 1;\n")
        f.write("        else hi = mid - 1;\n")
        f.write("    }\n")
        f.write("    return 0;\n")
        f.write("}\n\n")
        f.write("/* Layer extractors */\n")
        f.write("static inline uint32_t cnbe_get_radical(uint32_t c)   { return (c>>24)&0xFF; }\n")
        f.write("static inline uint32_t cnbe_get_strokes(uint32_t c)   { return (c>>19)&0x1F; }\n")
        f.write("static inline uint32_t cnbe_get_struct(uint32_t c)    { return (c>>15)&0xF;  }\n")
        f.write("static inline uint32_t cnbe_get_index(uint32_t c)     { return (c>>4)&0x7FF; }\n")
        f.write("static inline uint32_t cnbe_get_ext(uint32_t c)       { return c&0xF;        }\n")
        f.write("#endif\n")
    sz = os.path.getsize(os.path.join(OUT, "cnbe_table.h"))
    print(f"  ✓ cnbe_table.h (sorted, {sz//1024} KB)")

    # 写入排序后的 unicode_table.h
    with open(os.path.join(OUT, "unicode_table.h"), "w", encoding="utf-8") as f:
        f.write("// Unicode table (Sorted by Unicode value)\n")
        f.write("#ifndef UNICODE_TABLE_H\n#define UNICODE_TABLE_H\n#include <stdint.h>\n\n")
        f.write("static const uint32_t unicode_table[8105] = {\n")
        for i, ucp in enumerate(unicodes):
            f.write(f"    0x{ucp:04X}, // [{i:4d}] '{chars[i]}'\n")
        f.write("};\n\n")
        f.write("/* Hash lookup helper: unicode & 0x7FFF → index */\n")
        f.write("extern const int16_t cnbe_hash_index[32768][2];\n")
        f.write("static inline uint32_t cnbe_enc_hash(uint32_t ucp) {\n")
        f.write("    int h = ucp & 0x7FFF;\n")
        f.write("    int idx;\n")
        f.write("    idx = cnbe_hash_index[h][0];\n")
        f.write("    if (idx >= 0 && unicode_table[idx] == ucp) return cnbe_table[idx];\n")
        f.write("    idx = cnbe_hash_index[h][1];\n")
        f.write("    if (idx >= 0 && unicode_table[idx] == ucp) return cnbe_table[idx];\n")
        f.write("    return 0;\n")
        f.write("}\n#endif\n")
    print(f"  ✓ unicode_table.h (sorted, {os.path.getsize(os.path.join(OUT,'unicode_table.h'))//1024} KB)")

    # ============================================================
    # 2. 生成 15bit 哈希索引表
    # ============================================================
    from collections import defaultdict
    buckets = defaultdict(list)
    for i, u in enumerate(unicodes):
        buckets[u & 0x7FFF].append(i)

    max_collision = max(len(v) for v in buckets.values())
    total_buckets = len(buckets)
    collisions_1 = sum(1 for v in buckets.values() if len(v) == 1)
    collisions_2 = sum(1 for v in buckets.values() if len(v) == 2)
    print("\n哈希分析 (Unicode低15bit):")
    print(f"  总桶数: {total_buckets}, 最大碰撞: {max_collision}")
    print(f"  单条目桶: {collisions_1} ({100*collisions_1/total_buckets:.1f}%)")
    print(f"  双条目桶: {collisions_2} ({100*collisions_2/total_buckets:.1f}%)")

    with open(os.path.join(OUT, "cnbe_hash_index.h"), "w", encoding="utf-8") as f:
        f.write("// CNBE Hash Index Table — O(1) Unicode → CNBE lookup\n")
        f.write("// Hash: unicode & 0x7FFF (low 15 bits)\n")
        f.write("// Max collision: 2 entries per bucket\n")
        f.write("// Size: 32768 × 2 × 2 bytes = 131 KB\n")
        f.write("#include <stdint.h>\n\n")
        f.write("static const int16_t cnbe_hash_index[32768][2] = {\n")
        for h in range(32768):
            entries = buckets.get(h, [])
            e0 = entries[0] if len(entries) > 0 else -1
            e1 = entries[1] if len(entries) > 1 else -1
            if e0 >= 0 or e1 >= 0:
                u0 = unicodes[e0] if e0 >= 0 else 0
                c0 = cnbe_codes[e0] if e0 >= 0 else 0
                u1 = unicodes[e1] if e1 >= 0 else 0
                c1 = cnbe_codes[e1] if e1 >= 0 else 0
                f.write(f"    {{ {e0:4d}, {e1:4d} }}, // hash=0x{h:04X} U+{u0:04X}(0x{c0:08X})"
                        f"{' U+'+hex(u1)[2:]+'(0x'+hex(c1)[2:]+')' if e1>=0 else ''}\n")
            else:
                f.write(f"    {{ -1, -1 }}, // hash=0x{h:04X} (empty)\n")
        f.write("};\n")
    hash_sz = os.path.getsize(os.path.join(OUT, "cnbe_hash_index.h"))
    print(f"  ✓ cnbe_hash_index.h ({hash_sz//1024} KB, 32768 entries)")

    # 验证哈希正确性
    print("\n哈希正确性验证:")
    errors = 0
    for i, u in enumerate(unicodes):
        h = u & 0x7FFF
        found = False
        for col in range(2):
            idx = buckets[h][col] if col < len(buckets[h]) else -1
            if idx >= 0 and unicodes[idx] == u:
                found = True
                break
        if not found:
            errors += 1
    print(f"  验证 {len(unicodes)} 次查找, 错误: {errors}")
    print("  ✓ 100% 查找正确" if errors == 0 else f"  ✗ {errors} 错误")

    # 加速比总结
    n = len(unicodes)
    print(f"\n{'='*60}")
    print("查表优化总结")
    print(f"{'='*60}")
    print(f"  线性搜索:      {n//2:>7} 比较/次, 总 {n*n//2:>12,} 周期")
    print(f"  二分查找:      {math.ceil(math.log2(n)):>7} 比较/次, 总 {n*math.ceil(math.log2(n)):>12,} 周期")
    print(f"  哈希表:        1-2 比较/次, 总 {n:>12,} 周期")
    print(f"  二分加速比:    ~{n//2//math.ceil(math.log2(n)):>5}x")
    print(f"  哈希加速比:    ~{n//2:>5}x")
    print("\n完成！")

if __name__ == "__main__":
    main()
