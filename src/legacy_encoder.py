import zipfile, os, json
from collections import defaultdict

RADICAL_STROKES = {
    1:1,2:1,3:1,4:1,5:1,6:1,7:2,8:2,9:2,10:2,11:2,12:2,
    13:2,14:2,15:2,16:2,17:2,18:2,19:2,20:2,21:3,22:3,23:3,24:3,
    25:2,26:2,27:2,28:3,29:3,30:3,31:3,32:3,33:3,34:3,35:3,36:3,
    37:3,38:3,39:3,40:3,41:3,42:3,43:3,44:3,45:3,46:3,47:3,48:3,
    49:3,50:3,51:3,52:3,53:3,54:3,55:3,56:3,57:4,58:4,59:4,60:4,
    61:4,62:4,63:4,64:4,65:4,66:4,67:4,68:4,69:4,70:4,71:4,72:4,
    73:4,74:4,75:4,76:4,77:4,78:4,79:4,80:4,81:4,82:4,83:4,84:4,
    85:4,86:4,87:4,88:4,89:4,90:4,91:4,92:5,93:5,94:5,95:5,96:5,
    97:5,98:5,99:5,100:5,101:5,102:5,103:5,104:5,105:5,106:5,107:5,108:5,
    109:5,110:5,111:5,112:5,113:5,114:5,115:5,116:5,117:5,118:6,119:6,120:6,
    121:6,122:6,123:6,124:6,125:6,126:6,127:6,128:6,129:6,130:6,131:6,132:6,
    133:6,134:6,135:6,136:6,137:6,138:6,139:6,140:6,141:6,142:6,143:6,144:6,
    145:7,146:7,147:7,148:7,149:7,150:7,151:7,152:7,153:7,154:7,155:7,156:7,
    157:7,158:7,159:7,160:7,161:7,162:7,163:7,164:7,165:7,166:8,167:8,168:8,
    169:8,170:8,171:8,172:8,173:8,174:8,175:8,176:8,177:8,178:9,179:9,180:9,
    181:9,182:9,183:9,184:9,185:9,186:9,187:9,188:9,189:9,190:10,191:10,192:10,
    193:10,194:10,195:10,196:10,197:10,198:11,199:11,200:11,201:11,202:11,203:11,204:11,
    205:11,206:12,207:12,208:12,209:12,210:12,211:13,212:13,213:13,214:17,
}

LEFT_RADICALS = {9,61,64,75,85,86,94,96,112,113,118,120,140,149,154,159,162,163,167,170,196,195,38,60,62,63,66,69,72,73,74,76,77,78,79,80,81,83,84,87,88,91,93,100,102,104,106,107,109,110,114,115,117,121,124,126,130,134,136,137,139,141,142,144,145,146,147,148,150,151,152,153,155,156,157,158,160,161,164,165,166,168,169,171,173,174,175,176,177,178,180,181,182,183,184,185,187,188,189,190,191,192,193,194,196,197,198,199,200,201,202,203,204,205,206,207,208,209,210,211,212,213,214}

TOP_RADICALS = {1,3,7,8,14,19,20,24,25,26,27,30,31,32,34,35,36,37,39,40,41,42,44,45,46,47,48,49,50,51,53,54,55,56,57,58,59,65,67,68,70,71,82,89,90,92,95,97,98,99,101,103,105,108,111,116,119,122,123,125,127,128,129,131,132,133,135,138,143,172,179,186}

# 半包围/全包围判定
HUBS = {
    # 半包围-左下
    0x8FB6:7, 0x5EC4:7,
    # 半包围-上三
    0x95E8:8, 0x5188:8, 0x5300:8,
    # 半包围-下三
    0x5301:9, 0x531A:9,
    # 全包围
    0x56E7:0xA,
}

# 已知特定结构字
KNOWN_STRUCT = {
    "班":2, "辨":2, "辩":2, "辫":2, "瓣":2, "掰":2,
    "鼻":4, "鼾":4, "意":4, "慧":4, "冀":4, "翼":4,
    "照":4, "热":4, "烈":4, "煎":4, "熬":4, "熊":4,
    "品":11, "鑫":11, "森":11, "淼":11, "焱":11, "垚":11,
    "犇":11, "猋":11, "畾":11, "磊":11, "灥":11, "皛":11,
    "骉":11, "羴":11, "麤":11, "龘":11, "鱻":11, "馫":11,
}

STRUCT_NAMES = {
    0:"独体结构",1:"左右结构",2:"左中右结构",3:"上下结构",4:"上中下结构",
    5:"半包围-左上",6:"半包围-右上",7:"半包围-左下",8:"半包围-上三",
    9:"半包围-下三",10:"全包围",11:"品字结构",12:"镶嵌结构",13:"对称结构",
    14:"重叠结构",15:"未分类"
}

print("=" * 60)
print("CNBE v2.0 - 中文原生二进制编码系统")
print("=" * 60)

print("\n[1/5] 加载汉字列表...")
cp = "C:\\Users\\zairk\\Downloads\\1-8105\u7EAF\u6C49\u5B57\uFF08\u6309\u987A\u5E8F\u6392\u5217\uFF09.txt"
with open(cp, "r", encoding="utf-8") as f:
    chars = [line.strip() for line in f if line.strip()]
print(f"  加载 {len(chars)} 字")

print("\n[2/5] 解析Unihan...")
unihan = {}
strokes_db = {}
with zipfile.ZipFile("C:\\tmp\\Unihan.zip") as z:
    with z.open("Unihan_IRGSources.txt") as f:
        for line in f:
            line = line.decode("utf-8", errors="replace").strip()
            if not line or line.startswith("#") or "kRSUnicode" not in line:
                continue
            parts = line.split("\t")
            if len(parts) < 3: continue
            raw = parts[2].strip().split()[0]
            raw = raw.replace("'", "")  # 去掉星号/引号
            try:
                rad, add = raw.split(".")
                rad = int(rad); add = int(add)
                cp_int = int(parts[0].strip()[2:], 16)
                unihan[chr(cp_int)] = (rad, add)
            except: pass
    
    # kTotalStrokes
    with z.open("Unihan_IRGSources.txt") as f:
        for line in f:
            line = line.decode("utf-8", errors="replace").strip()
            if not line or line.startswith("#") or "kTotalStrokes" not in line: continue
            parts = line.split("\t")
            if len(parts) < 3: continue
            try:
                cp_int = int(parts[0].strip()[2:], 16)
                st = int(parts[2].strip().split()[0])
                strokes_db[chr(cp_int)] = st
            except: pass

print(f"  部首数据: {len(unihan)} 字, 笔画数据: {len(strokes_db)} 字")

# ---- 排查未覆盖字 ----
missing = [c for c in chars if c not in unihan]
# Check if they have kRSUnicode but with different format
with zipfile.ZipFile("C:\\tmp\\Unihan.zip") as z:
    with z.open("Unihan_IRGSources.txt") as f:
        data = f.read().decode("utf-8", errors="replace")

for m in missing[:5]:
    hexcp = f"U+{ord(m):04X}"
    if hexcp in data:
        lines = [l for l in data.split("\n") if hexcp in l and "kRSUnicode" in l]
        print(f"  MISSING {hexcp} {m}: found {len(lines)} lines")
        for l in lines[:2]:
            print(f"    raw: {l[:120]}")

covered = sum(1 for c in chars if c in unihan)
print(f"\n  8105字覆盖: {covered}/{len(chars)} ({100*covered/len(chars):.1f}%)")

# If still missing, try harder parsing
if len(missing) > 0:
    print("  尝试二次解析...")
    # Read file as raw and try different parsing
    unihan2 = {}
    with zipfile.ZipFile("C:\\tmp\\Unihan.zip") as z:
        with z.open("Unihan_IRGSources.txt") as f:
            for line in f:
                line = line.decode("utf-8", errors="replace").strip()
                if "kRSUnicode" not in line or line.startswith("#"): continue
                parts = line.split("\t")
                if len(parts) < 3: continue
                raw = parts[2].strip().split()[0]
                # Try all possible formats
                for sep in ["'", "\\u2019", "\\u2018", "\u02BC"]:
                    if sep in raw:
                        raw = raw.replace(sep, "")
                try:
                    rad, add = raw.split(".")
                    rad = int(rad); add = int(add)
                    cp_int = int(parts[0].strip()[2:], 16)
                    unihan2[chr(cp_int)] = (rad, add)
                except: pass
    covered2 = sum(1 for c in chars if c in unihan2)
    print(f"  二次解析覆盖: {covered2}/{len(chars)}")
    if covered2 > covered:
        print(f"  改进: +{covered2-covered} 字")
        unihan = unihan2

def get_rad(char):
    if char not in unihan: return 0
    r = unihan[char][0]
    return r if 1 <= r <= 214 else 0

def get_strokes(char):
    if char in strokes_db: return strokes_db[char]
    if char in unihan:
        r, a = unihan[char]
        if 1 <= r <= 214: return RADICAL_STROKES.get(r, 0) + a
        return a
    return 0

def get_struct(char, rc, sc):
    if char in KNOWN_STRUCT: return KNOWN_STRUCT[char]
    # 检查包围结构
    for hub_cp, st in HUBS.items():
        hub_char = chr(hub_cp)
        if hub_char in char and char != hub_char: return st
    # 半包围-左上 (厂广尸户疒气)
    for hcp in [0x5382, 0x5E7F, 0x5C38, 0x6237, 0x7589, 0x6C14]:
        hc = chr(hcp)
        if hc in char and char != hc: return 5
    if rc in LEFT_RADICALS: return 1
    if rc in TOP_RADICALS: return 3
    return 0 if sc <= 5 else 1

print("\n[3/5] 编码...")
cdat = []
groups = defaultdict(list)
rad0_cnt = 0
for c in chars:
    rc = get_rad(c)
    sc = min(get_strokes(c), 31)
    if rc == 0:
        rad0_cnt += 1
        if sc == 0: sc = 8
    st = get_struct(c, rc, sc)
    cdat.append((c, rc, sc, st))
    groups[(rc, sc, st)].append(c)

for k in groups: groups[k].sort(key=lambda x: chars.index(x))
cix = {}
for k, cl in groups.items():
    for i, c in enumerate(cl): cix[c] = i

results = []
for c, rc, sc, st in cdat:
    ix = cix[c]
    code = (rc << 24) | (sc << 19) | (st << 15) | (ix << 4) | 0
    results.append({"char":c, "unicode":f"U+{ord(c):04X}", "code":code, "hex":f"0x{code:08X}", "bin":f"{code:032b}", "rc":rc, "sc":sc, "st":st, "idx":ix})

print(f"  编码 {len(results)} 字")
print(f"  部首未覆盖(rc=0): {rad0_cnt}, 笔画未覆盖(sc=0): {sum(1 for r in results if r['sc']==0)}")

print("\n[4/5] 验证...")
codes = [r["code"] for r in results]
dup_cnt = len(codes) - len(set(codes))
dup_map = defaultdict(list)
for r in results: dup_map[r["code"]].append(r["char"])
dups = {k:v for k,v in dup_map.items() if len(v) > 1}
print(f"  总: {len(codes)}, 唯一: {len(set(codes))}, 重复: {dup_cnt}")
if dup_cnt == 0: print("  [OK] 全唯一")
else:
    for k,v in dups.items(): print(f"    DUP 0x{k:08X}: {','.join(v)}")

util = len(results) / 2**32 * 100
max_gs = max(len(v) for v in groups.values())
print(f"  空间利用率: {util:.6f}%, 分组: {len(groups)}, 最大组: {max_gs}/2048")

struct_cnt = defaultdict(int)
for r in results: struct_cnt[r["st"]] += 1
print("  结构分布:")
for sc in sorted(struct_cnt):
    print(f"    {STRUCT_NAMES[sc]}: {struct_cnt[sc]}")

rad_cnt = defaultdict(int)
for r in results: rad_cnt[r["rc"]] += 1
print("  部首Top10:")
for rc, cnt in sorted(rad_cnt.items(), key=lambda x: -x[1])[:10]:
    rn = "无" if rc==0 else f"#{rc}"
    print(f"    {rn}: {cnt}")

print("\n[5/5] Excel...")
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "CNBE完整编码表"

hf = Font(name="Arial", bold=True, size=11, color="FFFFFF")
hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
ha = Alignment(horizontal="center", vertical="center")
cf = Font(name="Arial", size=10)
ca = Alignment(horizontal="center", vertical="center")
tb = Border(left=Side(style="thin",color="D0D0D0"), right=Side(style="thin",color="D0D0D0"), top=Side(style="thin",color="D0D0D0"), bottom=Side(style="thin",color="D0D0D0"))

headers = ["序号","汉字","Unicode","CNBE(Hex)","CNBE(Dec)","CNBE(Bin)","部首区","笔画数","结构区","结构名称","字库区索引","扩展区"]
for i, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=i, value=h)
    c.font = hf; c.fill = hfill; c.alignment = ha; c.border = tb

for i, r in enumerate(results, 1):
    row = i + 1
    data = [i, r["char"], r["unicode"], r["hex"], r["code"], r["bin"], r["rc"], r["sc"], r["st"], STRUCT_NAMES[r["st"]], r["idx"], 0]
    for col, val in enumerate(data, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = cf; c.alignment = ca; c.border = tb

for col, w in enumerate([8,6,12,16,14,36,10,8,8,16,12,8], 1):
    ws.column_dimensions[chr(64+col)].width = w
ws.freeze_panes = "A2"

ws2 = wb.create_sheet("统计摘要")
for i, h in enumerate(["指标","数值"], 1):
    c = ws2.cell(row=1, column=i, value=h)
    c.font = hf; c.fill = hfill; c.alignment = ha; c.border = tb
stats = [
    ["编码方案","CNBE 32-bit"],
    ["汉字总数",str(len(results))],
    ["唯一编码数",str(len(set(codes)))],
    ["重复编码数",str(dup_cnt)],
    ["分组数(部首+笔画+结构)",str(len(groups))],
    ["最大组大小",f"{max_gs}/2048"],
    ["编码空间利用率",f"{util:.4f}%"],
    ["部首未覆盖(RC=0)",str(rad0_cnt)],
    ["位域分配","部首(8)+笔画(5)+结构(4)+字库(11)+扩展(4)"],
]
for i, (k,v) in enumerate(stats, 2):
    c1 = ws2.cell(row=i, column=1, value=k); c1.font=cf; c1.alignment=ca; c1.border=tb
    c2 = ws2.cell(row=i, column=2, value=v); c2.font=cf; c2.alignment=ca; c2.border=tb
ws2.column_dimensions["A"].width = 28; ws2.column_dimensions["B"].width = 28

ws3 = wb.create_sheet("结构分布")
for i, h in enumerate(["结构编码","结构名称","汉字数","占比"], 1):
    c = ws3.cell(row=1, column=i, value=h)
    c.font = hf; c.fill = hfill; c.alignment = ha; c.border = tb
for i, (sc, cnt) in enumerate(sorted(struct_cnt.items()), 2):
    data = [f"0x{sc:X}", STRUCT_NAMES[sc], cnt, f"{100*cnt/len(results):.1f}%"]
    for col, val in enumerate(data, 1):
        c = ws3.cell(row=i, column=col, value=val)
        c.font = cf; c.alignment = ca; c.border = tb
for col, w in zip(["A","B","C","D"],[12,18,12,10]):
    ws3.column_dimensions[col].width = w

ws4 = wb.create_sheet("笔画分布")
sc_grp = defaultdict(int)
for r in results: sc_grp[r["sc"]] += 1
for i, h in enumerate(["笔画数","汉字数","占比"], 1):
    c = ws4.cell(row=1, column=i, value=h)
    c.font = hf; c.fill = hfill; c.alignment = ha; c.border = tb
for i, (sc, cnt) in enumerate(sorted(sc_grp.items()), 2):
    data = [f"{sc}画", cnt, f"{100*cnt/len(results):.1f}%"]
    for col, val in enumerate(data, 1):
        c = ws4.cell(row=i, column=col, value=val)
        c.font = cf; c.alignment = ca; c.border = tb
for col, w in zip(["A","B","C"],[10,12,10]):
    ws4.column_dimensions[col].width = w

ws5 = wb.create_sheet("部首分布Top20")
for i, h in enumerate(["部首编码","部首编号","汉字数","占比"], 1):
    c = ws5.cell(row=1, column=i, value=h)
    c.font = hf; c.fill = hfill; c.alignment = ha; c.border = tb
for i, (rc, cnt) in enumerate(sorted(rad_cnt.items(), key=lambda x: -x[1])[:20], 2):
    rad_name = f"康熙#{rc}" if 1 <= rc <= 214 else ("无部首" if rc == 0 else f"扩展0x{rc:X}")
    data = [f"0x{rc:02X}", rad_name, cnt, f"{100*cnt/len(results):.1f}%"]
    for col, val in enumerate(data, 1):
        c = ws5.cell(row=i, column=col, value=val)
        c.font = cf; c.alignment = ca; c.border = tb
for col, w in zip(["A","B","C","D"],[12,14,12,10]):
    ws5.column_dimensions[col].width = w

out_dir = "C:\\Users\\zairk\\Documents\\Codex\\2026-06-22\\ni\\outputs"
os.makedirs(out_dir, exist_ok=True)
xlsx = os.path.join(out_dir, "CNBE编码目录.xlsx")
wb.save(xlsx)
print(f"  Excel: {xlsx}")

report = {
    "scheme":"CNBE 32-bit","version":"2.0",
    "bit_fields":"RC8+SC5+ST4+IX11+EX4",
    "total_chars":len(results), "unique_codes":len(set(codes)),
    "duplicates_count":dup_cnt, "groups_count":len(groups),
    "max_group_size":max_gs, "utilization_pct":util,
    "radical_uncovered":rad0_cnt,
    "structure_distribution":{STRUCT_NAMES[k]:v for k,v in sorted(struct_cnt.items())},
    "stroke_distribution":{str(k):v for k,v in sorted(sc_grp.items())},
    "top_radicals":{f"0x{rc:02X}":cnt for rc,cnt in sorted(rad_cnt.items(), key=lambda x:-x[1])[:10]},
    "encoding":{"bits_31_24":"radical_code","bits_23_19":"stroke_count","bits_18_15":"structure_type","bits_14_4":"index_in_group","bits_3_0":"extension_flags"}
}
rp = os.path.join(out_dir, "CNBE验证报告.json")
with open(rp, "w", encoding="utf-8") as f:
    json.dump(report, f, ensure_ascii=False, indent=2)
print(f"  报告: {rp}")

print("\n" + "=" * 60)
print("验证结论")
print("=" * 60)
ok = True
if dup_cnt == 0: print("  [OK] 全部编码唯一：通过")
else: print(f"  [FAIL] {dup_cnt} 组重复"); ok = False
if max_gs <= 2048: print(f"  [OK] 字库区容量充足 ({max_gs}/2048)")
else: print(f"  [FAIL] 字库区溢出 ({max_gs}>2048)"); ok = False
if rad0_cnt > 0: print(f"  [WARN] {rad0_cnt}字部首信息缺失")
else: print("  [OK] 全部字有部首")
print(f"\nExcel: {xlsx}")
print(f"Report: {rp}")
print("完成！")
